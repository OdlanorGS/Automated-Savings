"""
Modern Excel Formatter - Elegant Blue Theme
Clean, professional formatting with modern color palette
"""

import os
from tkinter import messagebox
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def apply_costa_formatting(file_path: str) -> bool:
    """
    Apply modern, elegant formatting with professional blue color scheme
    """

    if not os.path.exists(file_path):
        messagebox.showerror("Error", f"File not found:\n{file_path}")
        return False

    try:
        wb = load_workbook(file_path)
        ws = wb.worksheets[0]  # Main report sheet

        # Modern color palette (IBM Carbon-inspired)
        PRIMARY_BLUE = "0F62FE"     # Header color - modern blue
        SECONDARY_BLUE = "4589FF"   # Accent blue
        LIGHT_BLUE = "EDF5FF"       # Light blue tint
        WHITE = "FFFFFF"
        LIGHT_GRAY = "F8FAFB"       # Alternating rows - soft blue-gray
        TOTAL_GRAY = "E0E0E0"       # Total rows - neutral gray
        BORDER_GRAY = "D0D0D0"      # Subtle borders
        
        # Define modern, elegant styles
        header_font = Font(bold=True, color=WHITE, size=11, name='Segoe UI')
        header_fill = PatternFill(start_color=PRIMARY_BLUE, end_color=PRIMARY_BLUE, fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        title_font = Font(bold=True, size=12, color=PRIMARY_BLUE, name='Segoe UI')

        total_font = Font(bold=True, size=11, name='Segoe UI')
        total_fill = PatternFill(start_color=TOTAL_GRAY, end_color=TOTAL_GRAY, fill_type="solid")

        thin_border = Border(
            left=Side(style="thin", color=BORDER_GRAY),
            right=Side(style="thin", color=BORDER_GRAY),
            top=Side(style="thin", color=BORDER_GRAY),
            bottom=Side(style="thin", color=BORDER_GRAY)
        )

        thick_bottom_border = Border(
            left=Side(style="thin", color=BORDER_GRAY),
            right=Side(style="thin", color=BORDER_GRAY),
            top=Side(style="thin", color=BORDER_GRAY),
            bottom=Side(style="medium", color=PRIMARY_BLUE)
        )
        
        # Find section headers and data rows
        sections = []
        current_section = None
        
        for row_idx in range(1, ws.max_row + 1):
            first_cell = ws.cell(row=row_idx, column=1)
            
            # Identify headers (they follow empty rows or are first row)
            if row_idx == 1 or (row_idx > 1 and not ws.cell(row=row_idx-1, column=1).value):
                if first_cell.value:
                    # This is a header row
                    current_section = {
                        'header_row': row_idx,
                        'data_start': row_idx + 1,
                        'data_end': None
                    }
                    sections.append(current_section)
            
            # Find end of section (empty row or TOTAL row)
            elif current_section and not first_cell.value:
                if current_section['data_end'] is None:
                    current_section['data_end'] = row_idx - 1
        
        # Handle last section
        if sections and sections[-1]['data_end'] is None:
            sections[-1]['data_end'] = ws.max_row
        
        # Format each section
        for section in sections:
            # Format header row
            header_row = section['header_row']
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=header_row, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thick_bottom_border
            
            # Format data rows
            for row_idx in range(section['data_start'], section['data_end'] + 1):
                first_cell_value = ws.cell(row=row_idx, column=1).value
                
                # Check if it's a total row
                is_total_row = (first_cell_value and
                              ('TOTAL' in str(first_cell_value).upper() or
                               'Total' in str(first_cell_value)))

                # Determine row color (alternating for non-total rows)
                if is_total_row:
                    row_fill = total_fill
                    row_font = total_font
                    row_border = Border(
                        left=Side(style="thin", color=BORDER_GRAY),
                        right=Side(style="thin", color=BORDER_GRAY),
                        top=Side(style="medium", color=PRIMARY_BLUE),
                        bottom=Side(style="medium", color=PRIMARY_BLUE)
                    )
                else:
                    # Alternating row colors with modern soft tones
                    if (row_idx - section['data_start']) % 2 == 0:
                        row_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
                    else:
                        row_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
                    row_font = Font(size=10, name='Segoe UI')
                    row_border = thin_border
                
                # Apply formatting to entire row
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col)
                    
                    if not is_total_row or col > 1:  # Don't override "TOTAL" text formatting
                        cell.font = row_font
                    else:
                        cell.font = total_font
                    
                    cell.fill = row_fill
                    cell.border = row_border
                    
                    # Align numbers to right, text to left
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        # Format numbers
                        if 'Price' in str(ws.cell(header_row, col).value) or 'price' in str(ws.cell(header_row, col).value):
                            cell.number_format = '$#,##0.0000'
                        elif '%' in str(ws.cell(header_row, col).value):
                            cell.number_format = '0.00"%"'
                        else:
                            cell.number_format = '#,##0.00'
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 30)  # Cap at 30 to avoid too wide
            ws.column_dimensions[column_letter].width = max(adjusted_width, 10)  # Min width 10
        
        # Add section titles (without merging to avoid corruption)
        # Just make the first cell of empty rows bold with section names
        
        # Format Monthly Summary sheet if it exists
        if len(wb.worksheets) > 1:
            ms = wb.worksheets[1]  # Monthly Summary sheet
            
            # Format headers
            for col in range(1, ms.max_column + 1):
                cell = ms.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thick_bottom_border
            
            # Format data rows
            for row_idx in range(2, ms.max_row + 1):
                for col in range(1, ms.max_column + 1):
                    cell = ms.cell(row=row_idx, column=col)

                    # Alternating colors with modern soft tones
                    if row_idx % 2 == 0:
                        cell.fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")

                    cell.border = thin_border

                    # Format numbers
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        cell.font = Font(size=10, name='Segoe UI')
                        if col > 1:  # Not the month column
                            cell.number_format = '#,##0.00'
                    else:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.font = Font(bold=True, size=10, name='Segoe UI')
            
            # Auto-fit columns
            for column in ms.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 20)
                ms.column_dimensions[column_letter].width = max(adjusted_width, 10)
        
        # Set print settings for professional output
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = False

        # Add professional header/footer
        ws.oddHeader.center.text = "Savings Tracker Pro - Analytics Report"
        ws.oddFooter.left.text = "Confidential"
        ws.oddFooter.center.text = "Page &P of &N"
        ws.oddFooter.right.text = "&D"
        
        # Freeze panes for easier viewing
        ws.freeze_panes = "B2"
        
        # Save the file
        wb.save(file_path)
        wb.close()
        
        print(f"✅ Formatting applied successfully")
        return True
        
    except PermissionError:
        messagebox.showerror(
            "Permission Denied",
            f"Please close the Excel file before formatting:\n{file_path}"
        )
        return False
    
    except Exception as e:
        print(f"Formatting error: {e}")
        messagebox.showerror("Format Error", f"Could not apply formatting:\n{str(e)}")
        return False