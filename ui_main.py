import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import os
import json
import pandas as pd
from PIL import Image, ImageTk
import calendar
from openpyxl import load_workbook
import threading
from functools import lru_cache
import pickle
import hashlib

# Local modules
from engine import run_savings_analysis
from formatter import apply_costa_formatting
from config_manager import load_config, select_hist_file, select_output_folder

# Try calendar import (safe fallback)
try:
    from tkcalendar import Calendar
    CALENDAR_AVAILABLE = True
except ImportError:
    CALENDAR_AVAILABLE = False


# ========== HELPER FUNCTION ==========
def parse_vendor_filter(vendor_string):
    """
    Smart parser that handles both old (comma) and new (pipe) separators.
    Returns a list of vendor names or "All".
    """
    if not vendor_string or vendor_string == "All":
        return "All"
    
    if not isinstance(vendor_string, str):
        return vendor_string  # Already a list
    
    # New format: pipe-separated
    if " | " in vendor_string:
        vendors = [v.strip() for v in vendor_string.split(" | ")]
        return vendors if vendors else "All"
    
    # Old format: comma-separated
    if "," in vendor_string:
        parts = vendor_string.split(",")
        common_suffixes = ['Inc', 'Llc', 'Ltd', 'Corp', 'Co', 'Lp', 'Usa']
        
        if any(part.strip() in common_suffixes for part in parts):
            return [vendor_string.strip()]
        
        vendors = [v.strip() for v in parts]
        return vendors if vendors else "All"
    
    return [vendor_string.strip()]


class HistoricalDataCache:
    """Manages cached historical data to avoid repeated Excel loads"""
    
    def __init__(self):
        self.data = None
        self.file_path = None
        self.file_hash = None
        self.cache_file = "hist_cache.pkl"
    
    def get_file_hash(self, file_path):
        """Get hash of file to detect changes"""
        try:
            stat = os.stat(file_path)
            return hashlib.md5(f"{file_path}{stat.st_size}{stat.st_mtime}".encode()).hexdigest()
        except:
            return None
    
    def load_data(self, file_path, force_reload=False):
        """Load historical data with caching"""
        current_hash = self.get_file_hash(file_path)
        
        if not force_reload and self.data is not None and self.file_path == file_path and self.file_hash == current_hash:
            print("✅ Using cached historical data")
            return self.data
        
        print(f"🔄 Loading historical data from: {file_path}")
        
        try:
            if not force_reload and os.path.exists(self.cache_file):
                try:
                    with open(self.cache_file, 'rb') as f:
                        cache_data = pickle.load(f)
                        if cache_data['file_hash'] == current_hash:
                            print("✅ Loaded from pickle cache")
                            self.data = cache_data['data']
                            self.file_path = file_path
                            self.file_hash = current_hash
                            return self.data
                except Exception as e:
                    print(f"Cache load failed: {e}")
            
            print("Reading Excel file... (this may take a moment)")
            
            usecols = [
                "Sage & SAMII[Product Code]",
                "Sage & SAMII[Vendor Cleaned]",
                "Sage & SAMII[Receipt Date]",
                "[SumBase_RCVD_QTY]",
                "[SumBase_Unit_Cost]"
            ]
            
            data = pd.read_excel(
                file_path,
                usecols=usecols,
                dtype={
                    "Sage & SAMII[Product Code]": str,
                    "Sage & SAMII[Vendor Cleaned]": str
                }
            )
            
            print("Optimizing data types...")
            data["Sage & SAMII[Receipt Date]"] = pd.to_datetime(data["Sage & SAMII[Receipt Date]"], errors="coerce")
            data["Sage & SAMII[Product Code]"] = data["Sage & SAMII[Product Code]"].astype('category')
            data["Sage & SAMII[Vendor Cleaned]"] = data["Sage & SAMII[Vendor Cleaned]"].astype('category')
            data = data.dropna(subset=["Sage & SAMII[Product Code]", "Sage & SAMII[Receipt Date]"])
            
            print(f"✅ Loaded {len(data):,} records")
            
            self.data = data
            self.file_path = file_path
            self.file_hash = current_hash
            
            try:
                with open(self.cache_file, 'wb') as f:
                    pickle.dump({'file_hash': current_hash, 'data': data}, f)
                print("✅ Saved to pickle cache")
            except Exception as e:
                print(f"Could not save cache: {e}")
            
            return self.data
            
        except Exception as e:
            print(f"❌ Error loading historical data: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def clear_cache(self):
        """Clear cached data"""
        self.data = None
        self.file_path = None
        self.file_hash = None
        if os.path.exists(self.cache_file):
            os.remove(self.cache_file)
        print("✅ Cache cleared")


class ProjectManager:
    """Handles saving and loading project configurations"""
    def __init__(self, storage_dir="saved_projects"):
        self.storage_dir = storage_dir
        os.makedirs(storage_dir, exist_ok=True)
        self.projects_file = os.path.join(storage_dir, "projects.json")
        self._ensure_projects_file()

    def _ensure_projects_file(self):
        if not os.path.exists(self.projects_file):
            with open(self.projects_file, 'w') as f:
                json.dump({}, f)

    def save_project(self, project_name, months, entries):
        try:
            with open(self.projects_file, 'r') as f:
                projects = json.load(f)
            
            projects[project_name] = {
                "months": months,
                "entries": entries,
                "last_modified": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            
            with open(self.projects_file, 'w') as f:
                json.dump(projects, f, indent=2)
            return True
        except Exception as e:
            print(f"Error saving project: {e}")
            return False

    def load_project(self, project_name):
        try:
            with open(self.projects_file, 'r') as f:
                projects = json.load(f)
            return projects.get(project_name)
        except Exception as e:
            print(f"Error loading project: {e}")
            return None

    def list_projects(self):
        try:
            with open(self.projects_file, 'r') as f:
                projects = json.load(f)
            return sorted(projects.keys())
        except Exception:
            return []

    def delete_project(self, project_name):
        try:
            with open(self.projects_file, 'r') as f:
                projects = json.load(f)
            
            if project_name in projects:
                del projects[project_name]
                with open(self.projects_file, 'w') as f:
                    json.dump(projects, f, indent=2)
                return True
            return False
        except Exception as e:
            print(f"Error deleting project: {e}")
            return False

    def get_project_info(self, project_name):
        project = self.load_project(project_name)
        if project:
            return {
                "name": project_name,
                "months": project.get("months"),
                "entry_count": len(project.get("entries", [])),
                "last_modified": project.get("last_modified")
            }
        return None


class ModernStyle:
    """Modern, elegant color scheme and styling"""
    PRIMARY = "#0F62FE"        # Modern blue
    SECONDARY = "#4589FF"      # Light blue
    ACCENT = "#0353E9"         # Darker blue for hover
    LIGHT = "#F8FAFB"          # Light blue-gray background
    WHITE = "#FFFFFF"          # Pure white
    TEXT_DARK = "#161616"      # Almost black
    TEXT_LIGHT = "#525252"     # Medium gray
    BORDER = "#E0E0E0"         # Light border
    SUCCESS = "#24A148"        # Green for success states
    WARNING = "#E87500"        # Orange for warnings
    ERROR = "#DA1E28"          # Red for errors

    @staticmethod
    def configure_styles():
        """Configure modern, elegant UI styles"""
        style = ttk.Style()
        style.theme_use('clam')

        # Treeview with modern styling
        style.configure("Treeview",
                       background=ModernStyle.WHITE,
                       foreground=ModernStyle.TEXT_DARK,
                       fieldbackground=ModernStyle.WHITE,
                       borderwidth=0,
                       rowheight=28,
                       font=('Segoe UI', 10))
        style.configure("Treeview.Heading",
                       background=ModernStyle.PRIMARY,
                       foreground=ModernStyle.WHITE,
                       font=('Segoe UI', 10, 'bold'),
                       borderwidth=0,
                       relief='flat')
        style.map('Treeview.Heading',
                 background=[('active', ModernStyle.SECONDARY)])

        # Primary accent button - modern flat blue
        style.configure("Accent.TButton",
                       background=ModernStyle.PRIMARY,
                       foreground=ModernStyle.WHITE,
                       font=('Segoe UI', 10, 'bold'),
                       borderwidth=0,
                       focuscolor='none',
                       relief='flat',
                       padding=(15, 8))
        style.map("Accent.TButton",
                 background=[('active', ModernStyle.ACCENT),
                           ('pressed', ModernStyle.ACCENT)])

        # Standard button - light gray
        style.configure("TButton",
                       background='#E0E0E0',
                       foreground=ModernStyle.TEXT_DARK,
                       font=('Segoe UI', 10),
                       borderwidth=0,
                       focuscolor='none',
                       relief='flat',
                       padding=(15, 8))
        style.map("TButton",
                 background=[('active', '#D0D0D0'),
                           ('pressed', '#C0C0C0')])

        # Frames and containers
        style.configure("TFrame", background=ModernStyle.LIGHT)

        # Card-like label frames
        style.configure("TLabelframe",
                       background=ModernStyle.WHITE,
                       foreground=ModernStyle.TEXT_DARK,
                       borderwidth=1,
                       relief="solid",
                       bordercolor=ModernStyle.BORDER)
        style.configure("TLabelframe.Label",
                       background=ModernStyle.WHITE,
                       foreground=ModernStyle.PRIMARY,
                       font=('Segoe UI', 11, 'bold'))

        # Labels
        style.configure("TLabel",
                       background=ModernStyle.WHITE,
                       foreground=ModernStyle.TEXT_DARK,
                       font=('Segoe UI', 10))
        style.configure("Header.TLabel",
                       font=('Segoe UI', 12, 'bold'),
                       foreground=ModernStyle.PRIMARY)

        # Entry fields
        style.configure("TEntry",
                       fieldbackground=ModernStyle.WHITE,
                       foreground=ModernStyle.TEXT_DARK,
                       borderwidth=1,
                       relief='solid')


class ReportUpdater:
    """Updates existing savings reports with fresh data"""
    
    def __init__(self, parent, config):
        self.parent = parent
        self.config = config
        self.report_path = None
        self.hist_path = config.get("hist_file", "")
    
    def show_updater_dialog(self):
        if not self.hist_path or not os.path.exists(self.hist_path):
            messagebox.showerror(
                "Missing Historical File",
                "Please set the Historical Data file in Settings before updating reports."
            )
            return
        
        dialog = tk.Toplevel(self.parent)
        dialog.title("Update Existing Report")
        dialog.geometry("700x450")
        dialog.configure(bg="#F5F9F3")
        dialog.grab_set()
        dialog.transient(self.parent)
        
        header = tk.Frame(dialog, bg="#2D5016", height=70)
        header.pack(fill="x")
        header.pack_propagate(False)
        
        tk.Label(header, text="🔄 Report Updater",
                font=('Segoe UI', 18, 'bold'),
                bg="#2D5016", fg="white").pack(pady=20)
        
        content = tk.Frame(dialog, bg="#F5F9F3", padx=30, pady=20)
        content.pack(fill="both", expand=True)
        
        instructions = tk.Label(
            content,
            text="Update an existing savings report with the latest data from your historical file.\n"
                 "This will refresh all quantities and savings without recreating the report.",
            bg="#F5F9F3",
            font=('Segoe UI', 10),
            justify="left",
            wraplength=620
        )
        instructions.pack(pady=(0, 20))
        
        file_frame = tk.LabelFrame(content, text="Select Report to Update",
                                  bg="#F5F9F3", font=('Segoe UI', 10, 'bold'),
                                  padx=15, pady=15)
        file_frame.pack(fill="x", pady=10)
        
        self.file_label = tk.Label(file_frame, text="No file selected",
                                   bg="#F5F9F3", fg="gray",
                                   font=('Segoe UI', 9))
        self.file_label.pack(pady=5)
        
        tk.Button(file_frame, text="📂 Browse for Report",
                 command=lambda: self.select_report(dialog),
                 bg="#6B9F3E", fg="white",
                 font=('Segoe UI', 10, 'bold'),
                 padx=20, pady=8,
                 cursor="hand2").pack(pady=5)
        
        options_frame = tk.LabelFrame(content, text="Update Options",
                                     bg="#F5F9F3", font=('Segoe UI', 10, 'bold'),
                                     padx=15, pady=15)
        options_frame.pack(fill="x", pady=10)
        
        self.update_all_var = tk.BooleanVar(value=True)
        self.update_empty_var = tk.BooleanVar(value=False)
        
        tk.Radiobutton(options_frame, text="Update ALL months (recommended)",
                      variable=self.update_all_var, value=True,
                      bg="#F5F9F3", font=('Segoe UI', 9),
                      command=lambda: self.update_empty_var.set(False)).pack(anchor="w", pady=2)
        
        tk.Radiobutton(options_frame, text="Update only empty months",
                      variable=self.update_empty_var, value=True,
                      bg="#F5F9F3", font=('Segoe UI', 9),
                      command=lambda: self.update_all_var.set(False)).pack(anchor="w", pady=2)
        
        self.status_label = tk.Label(content, text="Ready to update",
                                     bg="#F5F9F3", fg="#2D5016",
                                     font=('Segoe UI', 9, 'italic'))
        self.status_label.pack(pady=10)
        
        btn_frame = tk.Frame(content, bg="#F5F9F3")
        btn_frame.pack(pady=10)
        
        tk.Button(btn_frame, text="▶️ Update Report",
                 command=lambda: self.run_update(dialog),
                 bg="#2D5016", fg="white",
                 font=('Segoe UI', 11, 'bold'),
                 padx=30, pady=10,
                 cursor="hand2").pack(side="left", padx=5)
        
        tk.Button(btn_frame, text="Cancel",
                 command=dialog.destroy,
                 bg="#999999", fg="white",
                 font=('Segoe UI', 10),
                 padx=20, pady=10,
                 cursor="hand2").pack(side="left", padx=5)
    
    def select_report(self, dialog):
        path = filedialog.askopenfilename(
            title="Select Savings Report to Update",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            parent=dialog
        )
        
        if path:
            self.report_path = path
            self.file_label.config(text=os.path.basename(path), fg="black")
            self.status_label.config(text=f"Report loaded: {os.path.basename(path)}")
    
    def run_update(self, dialog):
        if not self.report_path:
            messagebox.showerror("No File", "Please select a report to update.", parent=dialog)
            return
        
        self.status_label.config(text="Updating report...")
        dialog.update()
        
        try:
            success = self.update_report(self.update_all_var.get())
            
            if success:
                messagebox.showinfo(
                    "Update Complete",
                    "Report updated successfully!\n\nUpdated file saved with '_Updated' suffix.",
                    parent=dialog
                )
                dialog.destroy()
            else:
                self.status_label.config(text="Update failed - see error message")
                
        except Exception as e:
            messagebox.showerror("Update Error", f"Failed to update report:\n{str(e)}", parent=dialog)
            self.status_label.config(text="Update failed")
    
    def update_report(self, update_all=True):
        print(f"\n=== Starting Report Update ===")
        print(f"Report: {self.report_path}")
        print(f"Mode: {'Update ALL months' if update_all else 'Update EMPTY months only'}")
        
        wb = load_workbook(self.report_path)
        
        if "Savings Report" not in wb.sheetnames:
            messagebox.showerror("Invalid Report", "This doesn't appear to be a valid savings report.")
            return False
        
        ws = wb["Savings Report"]
        
        hist = pd.read_excel(self.hist_path, dtype={"Sage & SAMII[Product Code]": str, "Sage & SAMII[Vendor Cleaned]": str})
        hist = hist.rename(columns={
            "Sage & SAMII[Product Code]": "ProductCode",
            "[SumBase_RCVD_QTY]": "BaseRcvdQty",
            "[SumBase_Unit_Cost]": "UnitCost",
            "Sage & SAMII[Receipt Date]": "ReceiptDate",
            "Sage & SAMII[Vendor Cleaned]": "Vendor"
        })
        hist = hist.dropna(subset=["Vendor", "ProductCode", "ReceiptDate", "UnitCost"]).copy()
        hist["ReceiptDate"] = pd.to_datetime(hist["ReceiptDate"], errors="coerce")
        hist = hist.dropna(subset=["ReceiptDate"])
        
        summary_header_row = None
        
        for row_idx in range(1, min(ws.max_row + 1, 50)):
            cell_value = ws.cell(row=row_idx, column=1).value
            
            if cell_value:
                cell_str = str(cell_value).strip()
                if cell_str == "Project" or cell_str == "Product Code":
                    summary_header_row = row_idx
                    print(f"Found SUMMARY header at row {row_idx}")
                    break
        
        if not summary_header_row:
            messagebox.showerror("Parse Error", 
                "Could not find SUMMARY column headers.\n"
                "Expected to find 'Project' or 'Product Code' in first column.")
            return False
        
        col_map = {}
        print(f"\nReading columns from row {summary_header_row}:")
        for col_idx in range(1, ws.max_column + 1):
            header_value = ws.cell(row=summary_header_row, column=col_idx).value
            if header_value:
                header_clean = str(header_value).strip()
                col_map[header_clean] = col_idx
                print(f"  Col {col_idx}: '{header_clean}'")
        
        required_cols = ["Product Code", "Old Price (Avg)", "New Price"]
        missing = [col for col in required_cols if col not in col_map]
        
        if missing:
            print(f"\nERROR: Missing columns: {missing}")
            print(f"Available columns: {list(col_map.keys())}")
            messagebox.showerror("Parse Error", 
                f"Missing required columns: {missing}\n\n"
                f"Found columns: {list(col_map.keys())}")
            return False
        
        print(f"\n✓ All required columns found")
        
        products = []
        data_start_row = summary_header_row + 1
        
        print(f"\nReading product data starting at row {data_start_row}:")
        for row_idx in range(data_start_row, ws.max_row + 1):
            first_cell = ws.cell(row=row_idx, column=1).value
            
            if not first_cell:
                continue
            if "QUANTITIES" in str(first_cell).upper():
                print(f"  Stopped at QUANTITIES section (row {row_idx})")
                break
            
            product_code = ws.cell(row=row_idx, column=col_map["Product Code"]).value
            compare_code = ws.cell(row=row_idx, column=col_map.get("Compare Code", col_map["Product Code"])).value
            old_price_val = ws.cell(row=row_idx, column=col_map["Old Price (Avg)"]).value
            new_price_val = ws.cell(row=row_idx, column=col_map["New Price"]).value
            
            if not product_code:
                continue
            
            try:
                old_price = float(old_price_val) if old_price_val else 0
                new_price = float(new_price_val) if new_price_val else 0
            except (ValueError, TypeError):
                print(f"  ⚠️ Skipping row {row_idx}: Invalid price values")
                continue
            
            print(f"  Row {row_idx}: {product_code} (Old: ${old_price:.4f}, New: ${new_price:.4f})")
            
            products.append({
                'product_code': str(product_code).strip(),
                'compare_code': str(compare_code or product_code).strip(),
                'old_price': old_price,
                'new_price': new_price
            })
        
        print(f"\n✓ Found {len(products)} products to update")
        
        if not products:
            messagebox.showwarning("No Products", "No products found to update.")
            return False
        
        qty_section_row = None
        for row_idx in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=1).value
            if cell_value and "QUANTITIES" in str(cell_value).upper():
                next_cell = ws.cell(row=row_idx + 1, column=1).value
                if next_cell and "Product" in str(next_cell):
                    qty_section_row = row_idx + 1
                    print(f"Found QUANTITIES header at row {qty_section_row}")
                    break
        
        save_section_row = None
        for row_idx in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=1).value
            if cell_value and "REALIZED" in str(cell_value).upper() and "SAVINGS" in str(cell_value).upper():
                next_cell = ws.cell(row=row_idx + 1, column=1).value
                if next_cell and "Product" in str(next_cell):
                    save_section_row = row_idx + 1
                    print(f"Found SAVINGS header at row {save_section_row}")
                    break
        
        if not qty_section_row or not save_section_row:
            messagebox.showerror("Parse Error", "Could not find QUANTITIES and SAVINGS sections.")
            return False
        
        month_cols = {}
        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(row=qty_section_row, column=col_idx).value
            if header:
                header_str = str(header).strip()
                if header_str not in ["Product", "Total Qty", "Total", ""]:
                    month_cols[col_idx] = header_str
        
        print(f"\n✓ Found {len(month_cols)} month columns: {list(month_cols.values())}")
        
        updates_made = 0
        
        for product_info in products:
            product = product_info['product_code']
            old_price = product_info['old_price']
            new_price = product_info['new_price']
            savings_per_unit = old_price - new_price
            
            print(f"\nUpdating: {product}")
            
            product_hist = hist[hist["ProductCode"] == product].copy()
            
            if product_hist.empty:
                print(f"  ⚠️ No historical data found")
                continue
            
            product_hist["YearMonth"] = product_hist["ReceiptDate"].dt.to_period("M")
            monthly_data = product_hist.groupby("YearMonth").agg(
                Qty=("BaseRcvdQty", "sum")
            ).reset_index()
            
            for col_idx, month_name in month_cols.items():
                try:
                    parts = month_name.split()
                    if len(parts) != 2:
                        continue
                        
                    month_abbr, year_short = parts
                    month_num = list(calendar.month_abbr).index(month_abbr.upper())
                    year = 2000 + int(year_short)
                    target_period = pd.Period(f"{year}-{month_num:02d}", freq="M")
                    
                except (ValueError, IndexError):
                    print(f"  ⚠️ Could not parse month: {month_name}")
                    continue
                
                should_update = update_all
                if not update_all:
                    qty_row = self.find_product_row(ws, product, qty_section_row)
                    if qty_row:
                        current_val = ws.cell(row=qty_row, column=col_idx).value
                        should_update = (current_val is None or current_val == 0 or current_val == "")
                
                if should_update:
                    month_data = monthly_data[monthly_data["YearMonth"] == target_period]
                    
                    if not month_data.empty:
                        qty = month_data.iloc[0]["Qty"]
                        savings = qty * savings_per_unit
                        
                        qty_row = self.find_product_row(ws, product, qty_section_row)
                        if qty_row:
                            ws.cell(row=qty_row, column=col_idx, value=round(qty, 2))
                            updates_made += 1
                        
                        save_row = self.find_product_row(ws, product, save_section_row)
                        if save_row:
                            ws.cell(row=save_row, column=col_idx, value=round(savings, 2))
                        
                        print(f"  ✓ {month_name}: Qty={qty:.0f}, Savings=${savings:.2f}")
                    else:
                        qty_row = self.find_product_row(ws, product, qty_section_row)
                        if qty_row:
                            ws.cell(row=qty_row, column=col_idx, value=0)
                        
                        save_row = self.find_product_row(ws, product, save_section_row)
                        if save_row:
                            ws.cell(row=save_row, column=col_idx, value=0)
        
        print("\n✓ Recalculating totals...")
        self.recalculate_totals(ws, month_cols, qty_section_row, save_section_row)
        
        if "Monthly Summary" in wb.sheetnames:
            print("✓ Updating Monthly Summary...")
            self.update_monthly_summary(wb, ws, month_cols, qty_section_row, save_section_row)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        base_name = os.path.splitext(self.report_path)[0]
        output_path = f"{base_name}_Updated_{timestamp}.xlsx"
        
        wb.save(output_path)
        wb.close()
        
        print(f"\n✅ Update complete: {updates_made} cells updated")
        print(f"✅ Saved to: {output_path}")
        
        return True
    
    def find_product_row(self, ws, product_code, section_start):
        for row_idx in range(section_start + 1, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=1).value
            
            if not cell_value or "TOTAL" in str(cell_value).upper():
                break
            
            if str(cell_value).strip() == str(product_code).strip():
                return row_idx
        
        return None
    
    def recalculate_totals(self, ws, month_cols, qty_start, save_start):
        qty_total_row = None
        save_total_row = None
        
        for row_idx in range(qty_start, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=1).value
            if cell_value and "TOTAL" in str(cell_value).upper():
                qty_total_row = row_idx
                break
        
        for row_idx in range(save_start, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=1).value
            if cell_value and "TOTAL" in str(cell_value).upper():
                save_total_row = row_idx
                break
        
        for col_idx in month_cols.keys():
            if qty_total_row:
                total = 0
                for row_idx in range(qty_start + 1, qty_total_row):
                    val = ws.cell(row=row_idx, column=col_idx).value
                    if isinstance(val, (int, float)):
                        total += val
                ws.cell(row=qty_total_row, column=col_idx, value=round(total, 2))
            
            if save_total_row:
                total = 0
                for row_idx in range(save_start + 1, save_total_row):
                    val = ws.cell(row=row_idx, column=col_idx).value
                    if isinstance(val, (int, float)):
                        total += val
                ws.cell(row=save_total_row, column=col_idx, value=round(total, 2))
        
        total_qty_col = None
        total_save_col = None
        
        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(row=qty_start, column=col_idx).value
            if header and "Total" in str(header):
                total_qty_col = col_idx
                break
        
        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(row=save_start, column=col_idx).value
            if header and "Total" in str(header):
                total_save_col = col_idx
                break
        
        if total_qty_col:
            for row_idx in range(qty_start + 1, qty_total_row if qty_total_row else ws.max_row):
                row_total = 0
                for col_idx in month_cols.keys():
                    val = ws.cell(row=row_idx, column=col_idx).value
                    if isinstance(val, (int, float)):
                        row_total += val
                ws.cell(row=row_idx, column=total_qty_col, value=round(row_total, 2))
        
        if total_save_col:
            for row_idx in range(save_start + 1, save_total_row if save_total_row else ws.max_row):
                row_total = 0
                for col_idx in month_cols.keys():
                    val = ws.cell(row=row_idx, column=col_idx).value
                    if isinstance(val, (int, float)):
                        row_total += val
                ws.cell(row=row_idx, column=total_save_col, value=round(row_total, 2))
    
    def update_monthly_summary(self, wb, ws, month_cols, qty_start, save_start):
        ws_summary = wb["Monthly Summary"]
        
        qty_total_row = None
        save_total_row = None
        
        for row_idx in range(qty_start, ws.max_row + 1):
            if "TOTAL" in str(ws.cell(row=row_idx, column=1).value or "").upper():
                qty_total_row = row_idx
                break
        
        for row_idx in range(save_start, ws.max_row + 1):
            if "TOTAL" in str(ws.cell(row=row_idx, column=1).value or "").upper():
                save_total_row = row_idx
                break
        
        for col_idx, month_name in month_cols.items():
            month_row = None
            for row_idx in range(2, ws_summary.max_row + 1):
                if ws_summary.cell(row=row_idx, column=1).value == month_name:
                    month_row = row_idx
                    break
            
            if not month_row:
                month_row = ws_summary.max_row + 1
                ws_summary.cell(row=month_row, column=1, value=month_name)
            
            qty_total = ws.cell(row=qty_total_row, column=col_idx).value if qty_total_row else 0
            save_total = ws.cell(row=save_total_row, column=col_idx).value if save_total_row else 0
            
            ws_summary.cell(row=month_row, column=2, value=round(qty_total or 0, 2))
            ws_summary.cell(row=month_row, column=3, value=round(save_total or 0, 2))


class SavingsApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Costa Farms - Savings Tracker Pro")
        self.root.geometry("1350x750")
        self.root.minsize(1150, 600)

        ModernStyle.configure_styles()
        self.root.configure(bg=ModernStyle.LIGHT)

        self.config = load_config()
        self.project_manager = ProjectManager()
        self.hist_cache = HistoricalDataCache()

        self.root.bind('<Control-s>', lambda e: self.save_current_project())
        self.root.bind('<Control-o>', lambda e: self.load_project_dialog())
        self.root.bind('<Control-n>', lambda e: self.new_project())
        self.root.bind('<Delete>', lambda e: self.delete_selected())

        self.create_menu_bar()
        self.create_header()
        self.create_main_frame()
        self.create_status_bar()

    def get_target_vendors(self, product_code, effective_date, months_to_track):
        try:
            if not self.config.get("hist_file") or not product_code:
                return []
            
            hist_path = self.config["hist_file"]
            hist = self.hist_cache.load_data(hist_path)
            
            if hist is None or hist.empty:
                print("No historical data available")
                return []
            
            eff_date = pd.to_datetime(effective_date)
            end_date = eff_date + pd.DateOffset(months=months_to_track)
            
            mask = (
                (hist["Sage & SAMII[Product Code]"] == product_code) &
                (hist['Sage & SAMII[Receipt Date]'] >= eff_date) &
                (hist['Sage & SAMII[Receipt Date]'] <= end_date)
            )
            
            filtered = hist[mask]
            vendors = filtered["Sage & SAMII[Vendor Cleaned]"].dropna().unique().tolist()
            vendors = sorted(vendors)
            
            return vendors
            
        except Exception as e:
            print(f"Error getting target vendors: {e}")
            import traceback
            traceback.print_exc()
            return []

    def show_report_updater(self):
        updater = ReportUpdater(self.root, self.config)
        updater.show_updater_dialog()

    def get_baseline_vendors(self, product_code, effective_date):
        try:
            if not self.config.get("hist_file") or not product_code:
                return []
            
            hist_path = self.config["hist_file"]
            hist = self.hist_cache.load_data(hist_path)
            
            if hist is None or hist.empty:
                print("No historical data available")
                return []
            
            eff_date = pd.to_datetime(effective_date)
            start_date = eff_date - pd.DateOffset(months=12)
            
            mask = (
                (hist["Sage & SAMII[Product Code]"] == product_code) &
                (hist['Sage & SAMII[Receipt Date]'] >= start_date) &
                (hist['Sage & SAMII[Receipt Date]'] < eff_date)
            )
            
            filtered = hist[mask]
            
            print(f"Debug Baseline: Found {len(filtered)} records for product '{product_code}'")
            
            vendors = filtered["Sage & SAMII[Vendor Cleaned]"].dropna().unique().tolist()
            vendors = sorted(vendors)
            
            print(f"Debug Baseline: Unique vendors: {vendors}")
            
            return vendors
            
        except Exception as e:
            print(f"Error getting baseline vendors: {e}")
            import traceback
            traceback.print_exc()
            return []

    def create_header(self):
        """Create modern, elegant header with blue gradient feel"""
        header = tk.Frame(self.root, bg=ModernStyle.PRIMARY, height=100)
        header.pack(fill="x", side="top")
        header.pack_propagate(False)

        # Left section with logo
        logo_frame = tk.Frame(header, bg=ModernStyle.PRIMARY)
        logo_frame.pack(side="left", padx=30, pady=15)

        logo_path = "costa_logo.png"
        if os.path.exists(logo_path):
            try:
                img = Image.open(logo_path)
                img = img.resize((70, 70), Image.Resampling.LANCZOS)
                photo = ImageTk.PhotoImage(img)
                logo_label = tk.Label(logo_frame, image=photo, bg=ModernStyle.PRIMARY)
                logo_label.image = photo
                logo_label.pack()
            except Exception as e:
                print(f"Could not load logo: {e}")
                self._create_logo_placeholder(logo_frame)
        else:
            self._create_logo_placeholder(logo_frame)

        # Center section with title
        title_frame = tk.Frame(header, bg=ModernStyle.PRIMARY)
        title_frame.pack(side="left", fill="both", expand=True, padx=15)

        tk.Label(title_frame, text="💰 Savings Tracker Pro",
                font=('Segoe UI', 24, 'bold'),
                bg=ModernStyle.PRIMARY,
                fg=ModernStyle.WHITE).pack(anchor="w", pady=(8, 0))
        tk.Label(title_frame, text="Global Sourcing Analytics & Reporting",
                font=('Segoe UI', 11),
                bg=ModernStyle.PRIMARY,
                fg="#B8D7FF").pack(anchor="w")

        # Right section with stats card
        stats_card = tk.Frame(header, bg="#FFFFFF", relief='flat')
        stats_card.pack(side="right", padx=30, pady=20)

        stats_inner = tk.Frame(stats_card, bg="#FFFFFF", padx=25, pady=12)
        stats_inner.pack()

        project_count = len(self.project_manager.list_projects())
        tk.Label(stats_inner, text=str(project_count),
                font=('Segoe UI', 22, 'bold'),
                bg="#FFFFFF",
                fg=ModernStyle.PRIMARY).pack()
        tk.Label(stats_inner, text="Active Projects",
                font=('Segoe UI', 9),
                bg="#FFFFFF",
                fg=ModernStyle.TEXT_LIGHT).pack()

    def _create_logo_placeholder(self, parent):
        """Create modern circular logo placeholder"""
        canvas = tk.Canvas(parent, width=70, height=70,
                          bg=ModernStyle.PRIMARY, highlightthickness=0)
        canvas.pack()
        # Outer circle
        canvas.create_oval(10, 10, 60, 60, fill="#FFFFFF", outline="")
        # Inner dollar sign
        canvas.create_text(35, 35, text="$", font=('Segoe UI', 28, 'bold'),
                          fill=ModernStyle.PRIMARY)

    def create_menu_bar(self):
        menubar = tk.Menu(self.root)

        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label="New Project Ctrl+N", command=self.new_project)
        file_menu.add_command(label="Save Project Ctrl+S", command=self.save_current_project)
        file_menu.add_command(label="Load Project... Ctrl+O", command=self.load_project_dialog)
        file_menu.add_separator()
        file_menu.add_command(label="📥 Import from Excel...", command=self.import_from_excel)
        file_menu.add_command(label="📄 Download Import Template...", command=self.download_import_template)
        file_menu.add_separator()
        file_menu.add_command(label="🔄 Update Existing Report...", command=self.show_report_updater)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.root.quit)
        menubar.add_cascade(label="File", menu=file_menu)

        project_menu = tk.Menu(menubar, tearoff=0)
        project_menu.add_command(label="Manage Projects...", command=self.manage_projects_dialog)
        project_menu.add_command(label="Export Project Data...", command=self.export_project_data)
        menubar.add_cascade(label="Project", menu=project_menu)

        settings_menu = tk.Menu(menubar, tearoff=0)
        settings_menu.add_command(label="Historical Data File...",
                                command=lambda: self.select_and_reload_hist_file())
        settings_menu.add_command(label="Output Folder...",
                                command=lambda: select_output_folder(self.config))
        settings_menu.add_separator()
        settings_menu.add_command(label="🔄 Reload Historical Data", command=self.reload_hist_cache)
        settings_menu.add_command(label="🗑️ Clear Data Cache", command=self.clear_hist_cache)
        settings_menu.add_separator()
        settings_menu.add_command(label="View Settings", command=self.show_current_settings)
        menubar.add_cascade(label="Settings", menu=settings_menu)

        help_menu = tk.Menu(menubar, tearoff=0)
        help_menu.add_command(label="Keyboard Shortcuts", command=self.show_shortcuts)
        help_menu.add_command(label="About", command=self.show_about)
        menubar.add_cascade(label="Help", menu=help_menu)

        self.root.config(menu=menubar)

    def create_main_frame(self):
        container = ttk.Frame(self.root, padding=20)
        container.pack(fill="both", expand=True)

        project_frame = ttk.LabelFrame(container, text=" Project Information ", padding=15)
        project_frame.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 15))
        project_frame.columnconfigure(1, weight=1)

        ttk.Label(project_frame, text="Project Name:").grid(row=0, column=0, sticky="w", padx=(0, 10))
        self.project_entry = ttk.Entry(project_frame, width=35, font=('Segoe UI', 10))
        self.project_entry.grid(row=0, column=1, sticky="ew", padx=(0, 10))
        
        ttk.Button(project_frame, text="📂 Load", command=self.load_project_dialog,
                  style="Accent.TButton", width=10).grid(row=0, column=2, padx=5)
        
        ttk.Label(project_frame, text="Months:").grid(row=0, column=3, sticky="e", padx=(20, 10))
        self.months_entry = ttk.Entry(project_frame, width=8, font=('Segoe UI', 10))
        self.months_entry.insert(0, "12")
        self.months_entry.grid(row=0, column=4, sticky="w")

        entries_frame = ttk.LabelFrame(container, text=" Product Entries ", padding=15)
        entries_frame.grid(row=1, column=0, columnspan=2, sticky="nsew", pady=(0, 15))
        container.rowconfigure(1, weight=1)

        tree_container = ttk.Frame(entries_frame)
        tree_container.pack(fill="both", expand=True)
        
        self.tree = ttk.Treeview(tree_container,
                                columns=("Product", "CompareCode", "Price", "Date", "TargetVendors", "BaselineVendors"),
                                show="headings", height=14)
        
        self.tree.heading("Product", text="Product Code")
        self.tree.heading("CompareCode", text="Compare Code")
        self.tree.heading("Price", text="New Price")
        self.tree.heading("Date", text="Effective Date")
        self.tree.heading("TargetVendors", text="Target Vendors (New)")
        self.tree.heading("BaselineVendors", text="Baseline Vendors (Compare)")
        
        self.tree.column("Product", width=130)
        self.tree.column("CompareCode", width=130)
        self.tree.column("Price", width=90)
        self.tree.column("Date", width=110)
        self.tree.column("TargetVendors", width=180)
        self.tree.column("BaselineVendors", width=180)

        self.tree.tag_configure('oddrow', background=ModernStyle.WHITE)
        self.tree.tag_configure('evenrow', background=ModernStyle.LIGHT)

        scrollbar = ttk.Scrollbar(tree_container, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)
        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        self.tree.bind('<Double-1>', lambda e: self.edit_selected())

        btn_frame = ttk.Frame(container)
        btn_frame.grid(row=2, column=0, columnspan=2, pady=(0, 10))

        buttons = [
            ("➕ Add Entry", self.add_entry_popup, "Accent.TButton"),
            ("📥 Import Excel", self.import_from_excel, "TButton"),
            ("✏️ Edit", self.edit_selected, "TButton"),
            ("🗑️ Delete", self.delete_selected, "TButton"),
            ("💾 Save Project", self.save_current_project, "TButton"),
            ("▶️ Run Analysis", self.run_tracker, "Accent.TButton")
        ]

        for text, command, style in buttons:
            ttk.Button(btn_frame, text=text, command=command,
                      style=style, width=16).pack(side="left", padx=5)

    def create_status_bar(self):
        """Create modern status bar with sleek design"""
        status = tk.Frame(self.root, bg=ModernStyle.PRIMARY, height=30)
        status.pack(fill="x", side="bottom")

        self.status_label = tk.Label(status, text="✓ Ready",
                                     bg=ModernStyle.PRIMARY,
                                     fg=ModernStyle.WHITE,
                                     font=('Segoe UI', 10),
                                     anchor="w", padx=15)
        self.status_label.pack(side="left", fill="x", expand=True)

        # Version label
        version_label = tk.Label(status, text="v2.0",
                                bg=ModernStyle.PRIMARY,
                                fg="#B8D7FF",
                                font=('Segoe UI', 9),
                                anchor="e", padx=15)
        version_label.pack(side="right")

    def update_status(self, message, icon="✓"):
        """Update status bar with optional icon"""
        self.status_label.config(text=f"{icon} {message}")
        self.root.update_idletasks()

    def new_project(self):
        if messagebox.askyesno("New Project", "Clear current project and start new?"):
            self.project_entry.delete(0, tk.END)
            self.months_entry.delete(0, tk.END)
            self.months_entry.insert(0, "12")
            for item in self.tree.get_children():
                self.tree.delete(item)
            self.update_status("New project created")

    def save_current_project(self):
        project_name = self.project_entry.get().strip()
        if not project_name:
            messagebox.showwarning("Missing Info", "Please enter a project name to save.")
            return

        months = self.months_entry.get().strip()
        if not months.isdigit():
            messagebox.showwarning("Invalid Input", "Months to track must be a number.")
            return

        entries = []
        for item in self.tree.get_children():
            vals = self.tree.item(item, "values")
            
            target_vendor_filter = parse_vendor_filter(vals[4] if len(vals) > 4 else "All")
            baseline_vendor_filter = parse_vendor_filter(vals[5] if len(vals) > 5 else "All")
            
            entries.append({
                "product_code": vals[0],
                "compare_code": vals[1],
                "new_price": float(vals[2]),
                "effective_date": str(vals[3]),
                "target_vendors": target_vendor_filter,
                "baseline_vendors": baseline_vendor_filter
            })

        if self.project_manager.save_project(project_name, int(months), entries):
            messagebox.showinfo("Success", f"✅ Project '{project_name}' saved successfully!")
            self.update_status(f"Saved: {project_name}")
            self.refresh_header()
        else:
            messagebox.showerror("Error", "Failed to save project.")

    def load_project_dialog(self):
        projects = self.project_manager.list_projects()
        if not projects:
            messagebox.showinfo("No Projects", "No saved projects found.")
            return

        dialog = tk.Toplevel(self.root)
        dialog.title("Load Project")
        dialog.geometry("600x500")
        dialog.configure(bg=ModernStyle.LIGHT)
        dialog.grab_set()
        dialog.transient(self.root)

        header = tk.Frame(dialog, bg=ModernStyle.PRIMARY, height=60)
        header.pack(fill="x")
        header.pack_propagate(False)
        tk.Label(header, text="Load Saved Project",
                font=('Segoe UI', 14, 'bold'),
                bg=ModernStyle.PRIMARY,
                fg=ModernStyle.WHITE).pack(pady=15)

        content = ttk.Frame(dialog, padding=20)
        content.pack(fill="both", expand=True)

        columns = ("Name", "Entries", "Months", "Modified")
        proj_tree = ttk.Treeview(content, columns=columns, show="headings", height=12)
        
        proj_tree.heading("Name", text="Project Name")
        proj_tree.heading("Entries", text="# Entries")
        proj_tree.heading("Months", text="Months")
        proj_tree.heading("Modified", text="Last Modified")
        
        proj_tree.column("Name", width=200)
        proj_tree.column("Entries", width=80, anchor="center")
        proj_tree.column("Months", width=80, anchor="center")
        proj_tree.column("Modified", width=180)

        scrollbar = ttk.Scrollbar(content, orient="vertical", command=proj_tree.yview)
        proj_tree.configure(yscroll=scrollbar.set)
        proj_tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        for idx, proj in enumerate(projects):
            info = self.project_manager.get_project_info(proj)
            tag = 'evenrow' if idx % 2 == 0 else 'oddrow'
            proj_tree.insert("", "end", values=(
                proj,
                info['entry_count'],
                info['months'],
                info['last_modified']
            ), tags=(tag,))

        def load_selected(event=None):
            selection = proj_tree.selection()
            if not selection:
                messagebox.showwarning("No Selection", "Please select a project to load.")
                return
            values = proj_tree.item(selection[0], "values")
            project_name = values[0]
            self.load_project(project_name)
            dialog.destroy()

        proj_tree.bind('<Double-1>', load_selected)

        btn_frame = ttk.Frame(dialog, padding=10)
        btn_frame.pack(fill="x")
        ttk.Button(btn_frame, text="Load", command=load_selected,
                  style="Accent.TButton", width=15).pack(side="right", padx=5)
        ttk.Button(btn_frame, text="Cancel", command=dialog.destroy,
                  width=15).pack(side="right", padx=5)

    def load_project(self, project_name):
        project_data = self.project_manager.load_project(project_name)
        if not project_data:
            messagebox.showerror("Error", f"Could not load project '{project_name}'.")
            return

        self.project_entry.delete(0, tk.END)
        self.project_entry.insert(0, project_name)
        self.months_entry.delete(0, tk.END)
        self.months_entry.insert(0, str(project_data["months"]))
        
        for item in self.tree.get_children():
            self.tree.delete(item)

        for idx, entry in enumerate(project_data["entries"]):
            tag = 'evenrow' if idx % 2 == 0 else 'oddrow'
            
            target_vendor_filter = entry.get("target_vendors", entry.get("vendors", "All"))
            if isinstance(target_vendor_filter, list):
                target_vendor_filter = " | ".join(target_vendor_filter)
            
            baseline_vendor_filter = entry.get("baseline_vendors", "All")
            if isinstance(baseline_vendor_filter, list):
                baseline_vendor_filter = " | ".join(baseline_vendor_filter)
            
            self.tree.insert("", "end", values=(
                entry["product_code"],
                entry["compare_code"],
                entry["new_price"],
                entry["effective_date"],
                target_vendor_filter,
                baseline_vendor_filter
            ), tags=(tag,))

        messagebox.showinfo("Success", f"✅ Project '{project_name}' loaded!")
        self.update_status(f"Loaded: {project_name}")

    def manage_projects_dialog(self):
        projects = self.project_manager.list_projects()
        if not projects:
            messagebox.showinfo("No Projects", "No saved projects found.")
            return

        dialog = tk.Toplevel(self.root)
        dialog.title("Manage Projects")
        dialog.geometry("700x500")
        dialog.configure(bg=ModernStyle.LIGHT)
        dialog.grab_set()

        header = tk.Frame(dialog, bg=ModernStyle.PRIMARY, height=60)
        header.pack(fill="x")
        header.pack_propagate(False)
        tk.Label(header, text="Manage Saved Projects",
                font=('Segoe UI', 14, 'bold'),
                bg=ModernStyle.PRIMARY,
                fg=ModernStyle.WHITE).pack(pady=15)

        content = ttk.Frame(dialog, padding=20)
        content.pack(fill="both", expand=True)

        columns = ("Name", "Entries", "Months", "Modified")
        proj_tree = ttk.Treeview(content, columns=columns, show="headings", height=14)
        
        for col in columns:
            proj_tree.heading(col, text=col)
        
        proj_tree.column("Name", width=200)
        proj_tree.column("Entries", width=100, anchor="center")
        proj_tree.column("Months", width=100, anchor="center")
        proj_tree.column("Modified", width=200)

        scrollbar = ttk.Scrollbar(content, orient="vertical", command=proj_tree.yview)
        proj_tree.configure(yscroll=scrollbar.set)
        proj_tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        for idx, proj in enumerate(projects):
            info = self.project_manager.get_project_info(proj)
            tag = 'evenrow' if idx % 2 == 0 else 'oddrow'
            proj_tree.insert("", "end", values=(
                proj,
                info['entry_count'],
                info['months'],
                info['last_modified']
            ), tags=(tag,))

        def delete_selected():
            selection = proj_tree.selection()
            if not selection:
                messagebox.showwarning("No Selection", "Please select a project to delete.")
                return
            
            values = proj_tree.item(selection[0], "values")
            proj_name = values[0]
            
            if messagebox.askyesno("Confirm Delete",
                                  f"Delete project '{proj_name}'?\n\nThis cannot be undone."):
                if self.project_manager.delete_project(proj_name):
                    proj_tree.delete(selection[0])
                    messagebox.showinfo("Success", f"✅ Project '{proj_name}' deleted.")
                    self.refresh_header()
                else:
                    messagebox.showerror("Error", "Failed to delete project.")

        btn_frame = ttk.Frame(dialog, padding=10)
        btn_frame.pack(fill="x")
        ttk.Button(btn_frame, text="Delete", command=delete_selected,
                  width=15).pack(side="right", padx=5)
        ttk.Button(btn_frame, text="Close", command=dialog.destroy,
                  style="Accent.TButton", width=15).pack(side="right", padx=5)

    def export_project_data(self):
        items = self.tree.get_children()
        if not items:
            messagebox.showwarning("No Data", "No entries to export.")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        
        if file_path:
            data = []
            for item in items:
                vals = self.tree.item(item, "values")
                data.append(vals)
            
            df = pd.DataFrame(data, columns=["Product Code", "Compare Code",
                                            "New Price", "Effective Date", 
                                            "Target Vendors", "Baseline Vendors"])
            df.to_csv(file_path, index=False)
            messagebox.showinfo("Success", f"✅ Data exported to {file_path}")

    def refresh_header(self):
        project_count = len(self.project_manager.list_projects())
        self.update_status(f"Total saved projects: {project_count}")

    def add_entry_popup(self, edit=False, item=None):
        popup = tk.Toplevel(self.root)
        popup.title("Add Product Entry" if not edit else "Edit Product Entry")
        popup.geometry("650x800")
        popup.configure(bg=ModernStyle.LIGHT)
        popup.grab_set()
        popup.transient(self.root)

        header = tk.Frame(popup, bg=ModernStyle.PRIMARY, height=50)
        header.pack(fill="x")
        header.pack_propagate(False)
        tk.Label(header, text="Product Entry" if not edit else "Edit Entry",
                font=('Segoe UI', 12, 'bold'),
                bg=ModernStyle.PRIMARY,
                fg=ModernStyle.WHITE).pack(pady=12)

        canvas = tk.Canvas(popup, bg=ModernStyle.LIGHT, highlightthickness=0)
        scrollbar = ttk.Scrollbar(popup, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        def on_mousewheel(event):
            if canvas.winfo_exists():
                canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        popup.bind("<MouseWheel>", on_mousewheel)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        content = ttk.Frame(scrollable_frame, padding=20)
        content.pack(fill="both", expand=True)

        ttk.Label(content, text="Product Code:").grid(row=0, column=0, pady=8, sticky="e", padx=(0, 10))
        product_entry = ttk.Entry(content, width=30, font=('Segoe UI', 10))
        product_entry.grid(row=0, column=1, columnspan=2, sticky="ew")

        use_compare_var = tk.BooleanVar(value=False)
        compare_entry = ttk.Entry(content, width=30, state="disabled", font=('Segoe UI', 10))

        def toggle_compare():
            if use_compare_var.get():
                compare_entry.config(state="normal")
            else:
                compare_entry.config(state="disabled")
                compare_entry.delete(0, tk.END)

        ttk.Checkbutton(content, text="Use different comparison code",
                       variable=use_compare_var,
                       command=toggle_compare).grid(row=1, column=0, columnspan=3, pady=8)

        ttk.Label(content, text="Compare Code:").grid(row=2, column=0, pady=8, sticky="e", padx=(0, 10))
        compare_entry.grid(row=2, column=1, columnspan=2, sticky="ew")
        
        note_label = ttk.Label(content, 
                              text="💡 Tip: Baseline vendors will be based on comparison code", 
                              font=('Segoe UI', 8, 'italic'),
                              foreground=ModernStyle.TEXT_LIGHT)
        note_label.grid(row=3, column=0, columnspan=3, sticky="w", pady=(0, 5))

        ttk.Label(content, text="New Price:").grid(row=4, column=0, pady=8, sticky="e", padx=(0, 10))
        price_entry = ttk.Entry(content, width=30, font=('Segoe UI', 10))
        price_entry.grid(row=4, column=1, columnspan=2, sticky="ew")

        ttk.Label(content, text="Effective Date:").grid(row=5, column=0, pady=8, sticky="e", padx=(0, 10))
        date_entry = ttk.Entry(content, width=20, font=('Segoe UI', 10))
        date_entry.insert(0, datetime.now().strftime("%Y-%m-%d"))
        date_entry.grid(row=5, column=1, sticky="ew")

        if CALENDAR_AVAILABLE:
            def open_calendar():
                cal_win = tk.Toplevel(popup)
                cal_win.title("Select Date")
                cal_win.configure(bg=ModernStyle.LIGHT)
                cal_win.transient(popup)
                
                cal = Calendar(cal_win, selectmode="day", date_pattern="yyyy-mm-dd")
                cal.pack(padx=10, pady=10)

                def set_date():
                    date_entry.delete(0, tk.END)
                    date_entry.insert(0, cal.get_date())
                    cal_win.destroy()

                ttk.Button(cal_win, text="Select", command=set_date,
                          style="Accent.TButton").pack(pady=5)

            ttk.Button(content, text="📅", width=4,
                      command=open_calendar).grid(row=5, column=2, sticky="w", padx=5)

        separator1 = ttk.Separator(content, orient='horizontal')
        separator1.grid(row=6, column=0, columnspan=3, sticky="ew", pady=(20, 10))
        
        target_header_frame = tk.Frame(content, bg=ModernStyle.ACCENT, height=30)
        target_header_frame.grid(row=7, column=0, columnspan=3, sticky="ew", pady=(0, 10))
        target_header_frame.grid_propagate(False)
        tk.Label(target_header_frame, text="🎯 TARGET VENDORS (New Product Code)", 
                font=('Segoe UI', 10, 'bold'),
                bg=ModernStyle.ACCENT, fg=ModernStyle.WHITE).pack(pady=5)
        
        ttk.Label(content, text="Vendors to track for received quantities on the NEW product:",
                 font=('Segoe UI', 9), foreground=ModernStyle.TEXT_LIGHT).grid(
                     row=8, column=0, columnspan=3, sticky="w", pady=(0, 2))
        
        target_date_label = ttk.Label(content, text="📅 Date range will show after refresh",
                                      font=('Segoe UI', 8, 'italic'),
                                      foreground=ModernStyle.SECONDARY)
        target_date_label.grid(row=9, column=0, columnspan=3, sticky="w", pady=(0, 2))
        
        ttk.Label(content, text="Uses: Product Code field above",
                 font=('Segoe UI', 7, 'italic'), foreground=ModernStyle.TEXT_LIGHT).grid(
                     row=10, column=0, columnspan=3, sticky="w", pady=(0, 5))
        
        target_vendor_frame = ttk.Frame(content)
        target_vendor_frame.grid(row=11, column=0, columnspan=3, sticky="ew", pady=5)

        target_vendor_mode = tk.StringVar(value="all")
        target_selected_vendors = []
        
        def target_mode_changed():
            if target_vendor_mode.get() == "specific":
                target_vendor_listbox.config(state="normal", bg="white")
            else:
                target_vendor_listbox.config(state="disabled", bg="#f0f0f0")
        
        def save_target_selection(event=None):
            nonlocal target_selected_vendors
            target_selected_vendors = [target_vendor_listbox.get(i) for i in target_vendor_listbox.curselection()]
        
        ttk.Radiobutton(target_vendor_frame, text="All Vendors",
                       variable=target_vendor_mode, value="all",
                       command=target_mode_changed).pack(anchor="w")
        ttk.Radiobutton(target_vendor_frame, text="Select Specific Vendors:",
                       variable=target_vendor_mode, value="specific",
                       command=target_mode_changed).pack(anchor="w", pady=(5, 0))

        target_list_frame = ttk.Frame(target_vendor_frame)
        target_list_frame.pack(fill="both", expand=True, padx=(20, 0), pady=5)
        
        target_vendor_listbox = tk.Listbox(target_list_frame, height=4, selectmode="multiple",
                                    font=('Segoe UI', 9), state="disabled", bg="#f0f0f0",
                                    exportselection=False)
        target_vendor_scrollbar = ttk.Scrollbar(target_list_frame, orient="vertical",
                                        command=target_vendor_listbox.yview)
        target_vendor_listbox.configure(yscrollcommand=target_vendor_scrollbar.set)
        target_vendor_listbox.bind('<<ListboxSelect>>', save_target_selection)
        
        target_vendor_listbox.pack(side="left", fill="both", expand=True)
        target_vendor_scrollbar.pack(side="right", fill="y")

        separator2 = ttk.Separator(content, orient='horizontal')
        separator2.grid(row=12, column=0, columnspan=3, sticky="ew", pady=(20, 10))
        
        baseline_header_frame = tk.Frame(content, bg=ModernStyle.SECONDARY, height=30)
        baseline_header_frame.grid(row=13, column=0, columnspan=3, sticky="ew", pady=(0, 10))
        baseline_header_frame.grid_propagate(False)
        tk.Label(baseline_header_frame, text="📊 BASELINE VENDORS (Comparison Code)", 
                font=('Segoe UI', 10, 'bold'),
                bg=ModernStyle.SECONDARY, fg=ModernStyle.WHITE).pack(pady=5)
        
        ttk.Label(content, text="Vendors to use for prior year cost calculation on COMPARISON code:",
                 font=('Segoe UI', 9), foreground=ModernStyle.TEXT_LIGHT).grid(
                     row=14, column=0, columnspan=3, sticky="w", pady=(0, 2))
        
        baseline_date_label = ttk.Label(content, text="📅 Date range will show after refresh",
                                       font=('Segoe UI', 8, 'italic'),
                                       foreground=ModernStyle.SECONDARY)
        baseline_date_label.grid(row=15, column=0, columnspan=3, sticky="w", pady=(0, 2))
        
        ttk.Label(content, text="Uses: Compare Code field above (or Product Code if not set)",
                 font=('Segoe UI', 7, 'italic'), foreground=ModernStyle.TEXT_LIGHT).grid(
                     row=16, column=0, columnspan=3, sticky="w", pady=(0, 5))
        
        baseline_vendor_frame = ttk.Frame(content)
        baseline_vendor_frame.grid(row=17, column=0, columnspan=3, sticky="ew", pady=5)

        baseline_vendor_mode = tk.StringVar(value="all")
        baseline_selected_vendors = []
        
        def baseline_mode_changed():
            if baseline_vendor_mode.get() == "specific":
                baseline_vendor_listbox.config(state="normal", bg="white")
            else:
                baseline_vendor_listbox.config(state="disabled", bg="#f0f0f0")
        
        def save_baseline_selection(event=None):
            nonlocal baseline_selected_vendors
            baseline_selected_vendors = [baseline_vendor_listbox.get(i) for i in baseline_vendor_listbox.curselection()]
        
        ttk.Radiobutton(baseline_vendor_frame, text="All Vendors",
                       variable=baseline_vendor_mode, value="all",
                       command=baseline_mode_changed).pack(anchor="w")
        ttk.Radiobutton(baseline_vendor_frame, text="Select Specific Vendors:",
                       variable=baseline_vendor_mode, value="specific",
                       command=baseline_mode_changed).pack(anchor="w", pady=(5, 0))

        baseline_list_frame = ttk.Frame(baseline_vendor_frame)
        baseline_list_frame.pack(fill="both", expand=True, padx=(20, 0), pady=5)
        
        baseline_vendor_listbox = tk.Listbox(baseline_list_frame, height=4, selectmode="multiple",
                                    font=('Segoe UI', 9), state="disabled", bg="#f0f0f0",
                                    exportselection=False)
        baseline_vendor_scrollbar = ttk.Scrollbar(baseline_list_frame, orient="vertical",
                                        command=baseline_vendor_listbox.yview)
        baseline_vendor_listbox.configure(yscrollcommand=baseline_vendor_scrollbar.set)
        baseline_vendor_listbox.bind('<<ListboxSelect>>', save_baseline_selection)
        
        baseline_vendor_listbox.pack(side="left", fill="both", expand=True)
        baseline_vendor_scrollbar.pack(side="right", fill="y")

        def load_vendors_for_product():
            refresh_btn.config(state="disabled", text="🔄 Loading...")
            popup.update()
            
            def load_in_background():
                try:
                    try:
                        months_to_track = int(self.months_entry.get().strip())
                    except:
                        months_to_track = 12
                    
                    try:
                        effective_date = date_entry.get().strip()
                    except:
                        effective_date = datetime.now().strftime("%Y-%m-%d")
                    
                    new_product_code = product_entry.get().strip()
                    target_vendors = []
                    if new_product_code and effective_date:
                        target_vendors = self.get_target_vendors(new_product_code, effective_date, months_to_track)
                    
                    if use_compare_var.get() and compare_entry.get().strip():
                        comparison_code = compare_entry.get().strip()
                    else:
                        comparison_code = new_product_code
                    
                    baseline_vendors = []
                    if comparison_code and effective_date:
                        baseline_vendors = self.get_baseline_vendors(comparison_code, effective_date)
                    
                    popup.after(0, lambda: update_vendor_ui(target_vendors, baseline_vendors, effective_date, months_to_track))
                    
                except Exception as e:
                    popup.after(0, lambda: show_error(str(e)))
            
            def update_vendor_ui(target_vendors, baseline_vendors, effective_date, months_to_track):
                try:
                    current_target_selections = target_selected_vendors.copy()
                    current_baseline_selections = baseline_selected_vendors.copy()
                    
                    target_vendor_listbox.config(state="normal")
                    target_vendor_listbox.delete(0, tk.END)
                    
                    if target_vendors:
                        for vendor in target_vendors:
                            target_vendor_listbox.insert(tk.END, vendor)
                        
                        for i, vendor in enumerate(target_vendors):
                            if vendor in current_target_selections:
                                target_vendor_listbox.selection_set(i)
                    else:
                        target_vendor_listbox.insert(tk.END, "(No vendors found in date range)")
                    
                    if target_vendor_mode.get() == "all":
                        target_vendor_listbox.config(state="disabled", bg="#f0f0f0")
                    else:
                        target_vendor_listbox.config(bg="white")
                    
                    eff_dt = pd.to_datetime(effective_date)
                    end_dt = eff_dt + pd.DateOffset(months=months_to_track)
                    target_date_label.config(text=f"📅 {eff_dt.strftime('%Y-%m-%d')} to {end_dt.strftime('%Y-%m-%d')} ({len(target_vendors)} vendors)")
                    
                    baseline_vendor_listbox.config(state="normal")
                    baseline_vendor_listbox.delete(0, tk.END)
                    
                    if baseline_vendors:
                        for vendor in baseline_vendors:
                            baseline_vendor_listbox.insert(tk.END, vendor)
                        
                        for i, vendor in enumerate(baseline_vendors):
                            if vendor in current_baseline_selections:
                                baseline_vendor_listbox.selection_set(i)
                    else:
                        baseline_vendor_listbox.insert(tk.END, "(No vendors found in date range)")
                    
                    if baseline_vendor_mode.get() == "all":
                        baseline_vendor_listbox.config(state="disabled", bg="#f0f0f0")
                    else:
                        baseline_vendor_listbox.config(bg="white")
                    
                    start_dt = eff_dt - pd.DateOffset(months=12)
                    code_used = "Compare Code" if use_compare_var.get() and compare_entry.get().strip() else "Product Code"
                    baseline_date_label.config(
                        text=f"📅 {start_dt.strftime('%Y-%m-%d')} to {eff_dt.strftime('%Y-%m-%d')} ({len(baseline_vendors)} vendors) - Using: {code_used}"
                    )
                    
                finally:
                    refresh_btn.config(state="normal", text="🔄 Refresh Vendor Lists")
            
            def show_error(error_msg):
                messagebox.showerror("Loading Error", f"Failed to load vendors:\n{error_msg}")
                refresh_btn.config(state="normal", text="🔄 Refresh Vendor Lists")
            
            thread = threading.Thread(target=load_in_background, daemon=True)
            thread.start()

        refresh_frame = ttk.Frame(content)
        refresh_frame.grid(row=18, column=0, columnspan=3, pady=10)

        refresh_btn = ttk.Button(refresh_frame, text="🔄 Refresh Vendor Lists",
                    command=load_vendors_for_product, style="Accent.TButton")
        refresh_btn.pack()

        ttk.Label(refresh_frame, 
                text="Note: First load may take a moment with large data files. Subsequent loads are cached.",
                font=('Segoe UI', 7, 'italic'),
                foreground=ModernStyle.TEXT_LIGHT).pack(pady=(5,0))

        if edit and item:
            vals = self.tree.item(item, "values")
            product_entry.insert(0, vals[0])
            
            compare_entry.config(state="normal")
            compare_entry.insert(0, vals[1])
            
            price_entry.insert(0, vals[2])
            date_entry.delete(0, tk.END)
            date_entry.insert(0, vals[3])
            
            popup.after(100, load_vendors_for_product)
            
            if len(vals) > 4 and vals[4] != "All":
                target_vendor_mode.set("specific")
                target_vendor_listbox.config(state="normal", bg="white")
                popup.after(200, lambda: self._select_vendors_in_listbox(target_vendor_listbox, vals[4]))
            
            if len(vals) > 5 and vals[5] != "All":
                baseline_vendor_mode.set("specific")
                baseline_vendor_listbox.config(state="normal", bg="white")
                popup.after(200, lambda: self._select_vendors_in_listbox(baseline_vendor_listbox, vals[5]))
            
            if vals[0] != vals[1]:
                use_compare_var.set(True)
            else:
                compare_entry.config(state="disabled")
        else:
            popup.after(100, load_vendors_for_product)

        def save_entry():
            try:
                product = product_entry.get().strip()
                if not product:
                    messagebox.showerror("Error", "Product code is required.")
                    return

                compare = compare_entry.get().strip() or product
                price = float(price_entry.get().strip())
                eff_date = datetime.strptime(date_entry.get().strip(), "%Y-%m-%d").date()

                if target_vendor_mode.get() == "all":
                    target_vendor_filter = "All"
                else:
                    target_vendor_listbox.config(state="normal")
                    selected_indices = target_vendor_listbox.curselection()
                    if not selected_indices:
                        messagebox.showerror("Error",
                                           "Please select at least one target vendor or choose 'All Vendors'.")
                        return
                    target_vendor_filter = [target_vendor_listbox.get(i) for i in selected_indices]
                    target_vendor_filter = [v for v in target_vendor_filter if not v.startswith("(No vendors")]
                    if not target_vendor_filter:
                        messagebox.showerror("Error",
                                           "No valid target vendors available in the date range.")
                        return
                    target_vendor_filter = " | ".join(target_vendor_filter)

                if baseline_vendor_mode.get() == "all":
                    baseline_vendor_filter = "All"
                else:
                    baseline_vendor_listbox.config(state="normal")
                    selected_indices = baseline_vendor_listbox.curselection()
                    if not selected_indices:
                        messagebox.showerror("Error",
                                           "Please select at least one baseline vendor or choose 'All Vendors'.")
                        return
                    baseline_vendor_filter = [baseline_vendor_listbox.get(i) for i in selected_indices]
                    baseline_vendor_filter = [v for v in baseline_vendor_filter if not v.startswith("(No vendors")]
                    if not baseline_vendor_filter:
                        messagebox.showerror("Error",
                                           "No valid baseline vendors available in the date range.")
                        return
                    baseline_vendor_filter = " | ".join(baseline_vendor_filter)

            except ValueError:
                messagebox.showerror("Error",
                                   "Please enter valid data.\nPrice must be numeric.\nDate format: YYYY-MM-DD")
                return

            if edit and item:
                self.tree.item(item, values=(product, compare, price, eff_date, 
                                            target_vendor_filter, baseline_vendor_filter))
                self.update_status(f"Updated: {product}")
            else:
                idx = len(self.tree.get_children())
                tag = 'evenrow' if idx % 2 == 0 else 'oddrow'
                self.tree.insert("", "end",
                               values=(product, compare, price, eff_date, 
                                      target_vendor_filter, baseline_vendor_filter),
                               tags=(tag,))
                self.update_status(f"Added: {product}")

            popup.destroy()

        def close_popup():
            popup.unbind("<MouseWheel>")
            popup.destroy()

        btn_frame = ttk.Frame(content)
        btn_frame.grid(row=19, column=0, columnspan=3, pady=20)
        
        ttk.Button(btn_frame, text="Save", command=save_entry,
                  style="Accent.TButton", width=15).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="Cancel", command=close_popup,
                  width=15).pack(side="left", padx=5)

        popup.protocol("WM_DELETE_WINDOW", close_popup)
        product_entry.focus()

    def _select_vendors_in_listbox(self, listbox, vendor_string):
        if isinstance(vendor_string, str) and vendor_string != "All":
            vendor_names = vendor_string.split(" | ")
            for i, vendor in enumerate(listbox.get(0, tk.END)):
                if vendor in vendor_names:
                    listbox.selection_set(i)
    
    def edit_selected(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("No Selection", "Please select an entry to edit.")
            return
        self.add_entry_popup(edit=True, item=selected[0])

    def delete_selected(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("No Selection", "Please select an entry to delete.")
            return

        if messagebox.askyesno("Confirm Delete", "Delete selected entry?"):
            vals = self.tree.item(selected[0], "values")
            self.tree.delete(selected[0])
            self.update_status(f"Deleted: {vals[0]}")
            
            for idx, item in enumerate(self.tree.get_children()):
                tag = 'evenrow' if idx % 2 == 0 else 'oddrow'
                self.tree.item(item, tags=(tag,))

    def run_tracker(self):
        project = self.project_entry.get().strip()
        months = self.months_entry.get().strip()

        if not project:
            messagebox.showwarning("Missing Info", "Please enter a project name.")
            return

        if not months.isdigit():
            messagebox.showwarning("Invalid Input", "Months to track must be a number.")
            return

        items = self.tree.get_children()
        if not items:
            messagebox.showwarning("No Entries", "Please add at least one entry before running.")
            return

        entries = []
        for item in items:
            vals = self.tree.item(item, "values")
            
            target_vendor_filter = parse_vendor_filter(vals[4] if len(vals) > 4 else "All")
            baseline_vendor_filter = parse_vendor_filter(vals[5] if len(vals) > 5 else "All")
            
            entries.append({
                "product_code": vals[0],
                "compare_code": vals[1],
                "new_price": float(vals[2]),
                "effective_date": pd.to_datetime(str(vals[3])).strftime("%Y-%m-%d"),
                "target_vendors": target_vendor_filter,
                "baseline_vendors": baseline_vendor_filter
            })

        self.update_status("Running analysis...")
        self.root.update_idletasks()

        try:
            file_path = run_savings_analysis(entries, project, int(months))
            if file_path:
                apply_costa_formatting(file_path)
                self.update_status(f"✅ Analysis complete: {project}")
        except Exception as e:
            messagebox.showerror("Error", f"Analysis failed:\n{str(e)}")
            self.update_status("Analysis failed")

    def import_from_excel(self):
        file_path = filedialog.askopenfilename(
            title="Select Excel File to Import",
            filetypes=[
                ("Excel files", "*.xlsx;*.xls"),
                ("All files", "*.*")
            ]
        )
        
        if not file_path:
            return
        
        try:
            df = pd.read_excel(file_path)
            
            required_columns = ["Product Code", "New Price", "Effective Date"]
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                messagebox.showerror(
                    "Invalid Template",
                    f"Missing required columns:\n{', '.join(missing_columns)}\n\n"
                    f"Required columns:\n"
                    f"- Product Code\n"
                    f"- New Price\n"
                    f"- Effective Date\n\n"
                    f"Optional columns:\n"
                    f"- Compare Code\n"
                    f"- Target Vendors\n"
                    f"- Baseline Vendors"
                )
                return
            
            imported = 0
            skipped = 0
            errors = []
            
            for idx, row in df.iterrows():
                try:
                    product_code = str(row["Product Code"]).strip()
                    
                    if not product_code or product_code.lower() == "nan":
                        skipped += 1
                        errors.append(f"Row {idx + 2}: Empty Product Code")
                        continue
                    
                    try:
                        new_price = float(row["New Price"])
                    except (ValueError, TypeError):
                        skipped += 1
                        errors.append(f"Row {idx + 2}: Invalid New Price '{row['New Price']}'")
                        continue
                    
                    try:
                        if isinstance(row["Effective Date"], str):
                            effective_date = pd.to_datetime(row["Effective Date"]).strftime("%Y-%m-%d")
                        else:
                            effective_date = pd.to_datetime(row["Effective Date"]).strftime("%Y-%m-%d")
                    except:
                        skipped += 1
                        errors.append(f"Row {idx + 2}: Invalid Effective Date '{row['Effective Date']}'")
                        continue
                    
                    compare_code = str(row.get("Compare Code", "")).strip()
                    if not compare_code or compare_code.lower() == "nan":
                        compare_code = product_code
                    
                    target_vendors = str(row.get("Target Vendors", "All")).strip()
                    if not target_vendors or target_vendors.lower() == "nan":
                        target_vendors = "All"
                    
                    baseline_vendors = str(row.get("Baseline Vendors", "All")).strip()
                    if not baseline_vendors or baseline_vendors.lower() == "nan":
                        baseline_vendors = "All"
                    
                    tree_idx = len(self.tree.get_children())
                    tag = 'evenrow' if tree_idx % 2 == 0 else 'oddrow'
                    
                    self.tree.insert("", "end", values=(
                        product_code,
                        compare_code,
                        new_price,
                        effective_date,
                        target_vendors,
                        baseline_vendors
                    ), tags=(tag,))
                    
                    imported += 1
                    
                except Exception as e:
                    skipped += 1
                    errors.append(f"Row {idx + 2}: {str(e)}")
            
            result_message = f"Import Complete!\n\n"
            result_message += f"✅ Successfully imported: {imported} entries\n"
            
            if skipped > 0:
                result_message += f"⚠️ Skipped: {skipped} entries\n\n"
                
                if errors:
                    result_message += "Errors:\n"
                    for error in errors[:10]:
                        result_message += f"  • {error}\n"
                    
                    if len(errors) > 10:
                        result_message += f"  ... and {len(errors) - 10} more\n"
            
            messagebox.showinfo("Import Results", result_message)
            self.update_status(f"Imported {imported} entries from Excel")
            
        except Exception as e:
            messagebox.showerror(
                "Import Error",
                f"Failed to import Excel file:\n{str(e)}\n\n"
                f"Please ensure the file is not open in Excel and follows the correct template format."
            )

    def download_import_template(self):
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile="Product_Import_Template.xlsx",
            title="Save Import Template")
        
        if not file_path:
            return
        
        try:
            template_data = {
                "Product Code": ["PROD001", "PROD002", "PROD003"],
                "Compare Code": ["PROD001", "OLDPROD002", ""],
                "New Price": [2.50, 3.75, 1.25],
                "Effective Date": ["2025-01-15", "2025-02-01", "2025-01-20"],
                "Target Vendors": ["All", "Vendor A | Vendor B", "Vendor X"],
                "Baseline Vendors": ["All", "Vendor C", "All"]
            }
            
            df = pd.DataFrame(template_data)
            
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name="Product Import", index=False)
                
                wb = writer.book
                ws = wb["Product Import"]
                
                instructions_data = {
                    "Column": [
                        "Product Code",
                        "Compare Code",
                        "New Price",
                        "Effective Date",
                        "Target Vendors",
                        "Baseline Vendors"
                    ],
                    "Required": ["YES", "NO", "YES", "YES", "NO", "NO"],
                    "Description": [
                        "The new product code to track quantities for (must match 'Sage & SAMII[Product Code]' in HistData)",
                        "Product code for baseline cost calculation (leave blank to use Product Code)",
                        "The new negotiated price",
                        "Date when new price becomes effective (format: YYYY-MM-DD)",
                        "Vendors to track for new product (use 'All' or list like 'Vendor A | Vendor B', must match 'Sage & SAMII[Vendor Cleaned]')",
                        "Vendors for baseline cost calculation (use 'All' or list like 'Vendor A | Vendor B')"
                    ],
                    "Example": [
                        "TRAY-001",
                        "TRAY-OLD-001",
                        "2.50",
                        "2025-01-15",
                        "China Vendor | Vietnam Supplier",
                        "All"
                    ]
                }
                
                instructions_df = pd.DataFrame(instructions_data)
                instructions_df.to_excel(writer, sheet_name="Instructions", index=False)
                
                from openpyxl.styles import Font, PatternFill, Alignment
                
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="2D5016", end_color="2D5016", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                ws_inst = wb["Instructions"]
                for cell in ws_inst[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="2D5016", end_color="2D5016", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                
                for column in ws_inst.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_inst.column_dimensions[column_letter].width = adjusted_width
            
            messagebox.showinfo(
                "Template Created",
                f"✅ Template saved successfully!\n\n"
                f"Location: {file_path}\n\n"
                f"The template includes:\n"
                f"• Sample data sheet\n"
                f"• Instructions sheet with column details\n\n"
                f"Fill in your product data and import using File > Import from Excel"
            )
            
            self.update_status("Template created successfully")
            
        except Exception as e:
            messagebox.showerror("Template Error", f"Failed to create template:\n{str(e)}")

    def select_and_reload_hist_file(self):
        old_path = self.config.get("hist_file", "")
        select_hist_file(self.config)
        new_path = self.config.get("hist_file", "")
        
        if new_path != old_path and new_path:
            self.reload_hist_cache()

    def reload_hist_cache(self):
        hist_path = self.config.get("hist_file", "")
        if not hist_path:
            messagebox.showwarning("No File", "Please set Historical Data file first.")
            return
        
        self.update_status("Reloading historical data...")
        self.root.update()
        
        try:
            def reload():
                self.hist_cache.load_data(hist_path, force_reload=True)
                self.root.after(0, lambda: self.update_status("✅ Historical data reloaded"))
                self.root.after(0, lambda: messagebox.showinfo("Success", "Historical data reloaded successfully!"))
            
            thread = threading.Thread(target=reload, daemon=True)
            thread.start()
            
        except Exception as e:
            messagebox.showerror("Reload Error", f"Failed to reload data:\n{str(e)}")
            self.update_status("Reload failed")

    def clear_hist_cache(self):
        if messagebox.askyesno("Clear Cache", "Clear cached historical data?\n\nData will be reloaded on next use."):
            self.hist_cache.clear_cache()
            messagebox.showinfo("Cache Cleared", "✅ Data cache cleared successfully!")
            self.update_status("Cache cleared")

    def show_current_settings(self):
        hist_file = self.config.get("hist_file", "Not set")
        out_folder = self.config.get("output_folder", "Not set")

        dialog = tk.Toplevel(self.root)
        dialog.title("Current Settings")
        dialog.geometry("550x300")
        dialog.configure(bg=ModernStyle.LIGHT)
        dialog.grab_set()
        dialog.transient(self.root)

        header = tk.Frame(dialog, bg=ModernStyle.PRIMARY, height=50)
        header.pack(fill="x")
        header.pack_propagate(False)
        tk.Label(header, text="Current Settings",
                font=('Segoe UI', 12, 'bold'),
                bg=ModernStyle.PRIMARY,
                fg=ModernStyle.WHITE).pack(pady=12)

        content = ttk.Frame(dialog, padding=20)
        content.pack(fill="both", expand=True)

        ttk.Label(content, text="Historical Data File:",
                font=('Segoe UI', 10, 'bold')).pack(anchor="w", pady=(0, 5))
        ttk.Label(content, text=hist_file, wraplength=500).pack(anchor="w", pady=(0, 15))

        ttk.Label(content, text="Output Folder:",
                font=('Segoe UI', 10, 'bold')).pack(anchor="w", pady=(0, 5))
        ttk.Label(content, text=out_folder, wraplength=500).pack(anchor="w", pady=(0, 15))

        ttk.Button(content, text="Close", command=dialog.destroy,
                style="Accent.TButton", width=15).pack(pady=10)

    def show_shortcuts(self):
        shortcuts = """
        Keyboard Shortcuts:
        
        Ctrl+N        New Project
        Ctrl+S        Save Project
        Ctrl+O        Load Project
        Delete        Delete Selected Entry
        Double-Click  Edit Entry
        """
        messagebox.showinfo("Keyboard Shortcuts", shortcuts)

    def show_about(self):
        dialog = tk.Toplevel(self.root)
        dialog.title("About")
        dialog.geometry("450x350")
        dialog.configure(bg=ModernStyle.LIGHT)
        dialog.grab_set()
        dialog.transient(self.root)
        dialog.resizable(False, False)

        header = tk.Frame(dialog, bg=ModernStyle.PRIMARY, height=100)
        header.pack(fill="x")
        header.pack_propagate(False)
        
        tk.Label(header, text="Savings Tracker Pro",
                font=('Segoe UI', 16, 'bold'),
                bg=ModernStyle.PRIMARY,
                fg=ModernStyle.WHITE).pack(pady=(20, 5))
        tk.Label(header, text="Version 3.9.6",
                font=('Segoe UI', 10),
                bg=ModernStyle.PRIMARY,
                fg=ModernStyle.LIGHT).pack()

        content = ttk.Frame(dialog, padding=30)
        content.pack(fill="both", expand=True)

        info_text = """
        Developed by: Ronaldo Garcia Sevilla
        Department: Global Sourcing @ Costa Farms

        Purpose: Advanced savings analysis and tracking
        for procurement optimization and cost reduction
        initiatives.

        Features:
        • Project-based savings tracking
        • Historical data analysis
        • Multi-product comparison
        • Excel bulk import capability
        • Date-based vendor filtering:
          - Target vendors (Product Code, forward period)
          - Baseline vendors (Compare Code, 12mo lookback)
        • Automated report generation
        • Report updating capability
        
        Data Requirements:
        Historical Excel file must contain columns:
        'Sage & SAMII[Product Code]'
        'Sage & SAMII[Vendor Cleaned]'
        'Sage & SAMII[Receipt Date]'
        '[SumBase_RCVD_QTY]'
        '[SumBase_Unit_Cost]'
        """

        tk.Label(content, text=info_text,
                font=('Segoe UI', 10),
                bg=ModernStyle.LIGHT,
                fg=ModernStyle.TEXT_DARK,
                justify="left").pack()

        ttk.Button(content, text="Close", command=dialog.destroy,
                  style="Accent.TButton", width=15).pack(pady=10)


if __name__ == "__main__":
    root = tk.Tk()
    app = SavingsApp(root)
    root.mainloop()