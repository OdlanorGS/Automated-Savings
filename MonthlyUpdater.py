"""
Report Updater - Add this to your ui_main.py

This module provides functionality to update existing savings reports
with the latest data without recreating the entire report.
"""

import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from tkinter import filedialog, messagebox
import tkinter as tk
from tkinter import ttk
import calendar


class ReportUpdater:
    """Updates existing savings reports with fresh data"""
    
    def __init__(self, parent, config):
        self.parent = parent
        self.config = config
        self.report_path = None
        self.hist_path = config.get("hist_file", "")
    
    def show_updater_dialog(self):
        """Show the report updater dialog with modern, elegant UI"""

        # Validate historical file is set
        if not self.hist_path or not os.path.exists(self.hist_path):
            messagebox.showerror(
                "Missing Historical File",
                "Please set the Historical Data file in Settings before updating reports."
            )
            return

        # Modern color palette
        colors = {
            'bg': '#F8FAFB',           # Light blue-gray background
            'card': '#FFFFFF',         # White cards
            'primary': '#0F62FE',      # Modern blue
            'primary_hover': '#0353E9',
            'secondary': '#4589FF',    # Light blue
            'success': '#24A148',      # Green
            'text': '#161616',         # Dark text
            'text_secondary': '#525252',
            'border': '#E0E0E0',
            'shadow': '#00000010'
        }

        # Create dialog
        dialog = tk.Toplevel(self.parent)
        dialog.title("Update Report")
        dialog.geometry("720x550")
        dialog.configure(bg=colors['bg'])
        dialog.grab_set()
        dialog.transient(self.parent)

        # Modern header with gradient effect
        header_frame = tk.Frame(dialog, bg=colors['primary'], height=100)
        header_frame.pack(fill="x")
        header_frame.pack_propagate(False)

        header_content = tk.Frame(header_frame, bg=colors['primary'])
        header_content.pack(expand=True)

        tk.Label(header_content, text="🔄",
                font=('Segoe UI', 32),
                bg=colors['primary'], fg="white").pack(pady=(5, 0))

        tk.Label(header_content, text="Report Updater",
                font=('Segoe UI', 22, 'bold'),
                bg=colors['primary'], fg="white").pack()

        tk.Label(header_content, text="Refresh your savings data with the latest information",
                font=('Segoe UI', 10),
                bg=colors['primary'], fg="#B8D7FF").pack(pady=(2, 5))

        # Main scrollable content area
        content = tk.Frame(dialog, bg=colors['bg'], padx=40, pady=25)
        content.pack(fill="both", expand=True)

        # File selection card
        file_card = tk.Frame(content, bg=colors['card'], relief='flat', bd=0)
        file_card.pack(fill="x", pady=(0, 15))

        # Card shadow effect
        file_card_inner = tk.Frame(file_card, bg=colors['card'], padx=25, pady=20)
        file_card_inner.pack(fill="x", padx=1, pady=1)

        tk.Label(file_card_inner, text="📁 Select Report File",
                font=('Segoe UI', 12, 'bold'),
                bg=colors['card'], fg=colors['text']).pack(anchor="w", pady=(0, 10))

        # File path display
        file_display_frame = tk.Frame(file_card_inner, bg='#F4F4F4', relief='flat')
        file_display_frame.pack(fill="x", pady=(0, 12))

        self.file_label = tk.Label(file_display_frame, text="No file selected",
                                   bg='#F4F4F4', fg=colors['text_secondary'],
                                   font=('Segoe UI', 10), anchor='w', padx=15, pady=12)
        self.file_label.pack(fill="x")

        # Browse button
        browse_btn = tk.Button(file_card_inner, text="Browse Files",
                              command=lambda: self.select_report(dialog),
                              bg=colors['secondary'], fg="white",
                              font=('Segoe UI', 10, 'bold'),
                              relief='flat', bd=0,
                              padx=25, pady=10,
                              cursor="hand2",
                              activebackground=colors['primary'])
        browse_btn.pack(anchor="w")

        # Options card
        options_card = tk.Frame(content, bg=colors['card'], relief='flat', bd=0)
        options_card.pack(fill="x", pady=(0, 15))

        options_card_inner = tk.Frame(options_card, bg=colors['card'], padx=25, pady=20)
        options_card_inner.pack(fill="x", padx=1, pady=1)

        tk.Label(options_card_inner, text="⚙️ Update Mode",
                font=('Segoe UI', 12, 'bold'),
                bg=colors['card'], fg=colors['text']).pack(anchor="w", pady=(0, 12))

        self.update_all_var = tk.BooleanVar(value=True)
        self.update_empty_var = tk.BooleanVar(value=False)

        # Radio button options
        radio_frame1 = tk.Frame(options_card_inner, bg=colors['card'])
        radio_frame1.pack(fill="x", pady=3)

        tk.Radiobutton(radio_frame1, text="Update ALL months",
                      variable=self.update_all_var, value=True,
                      bg=colors['card'], fg=colors['text'],
                      font=('Segoe UI', 10),
                      activebackground=colors['card'],
                      selectcolor=colors['card'],
                      command=lambda: self.update_empty_var.set(False)).pack(side="left")

        tk.Label(radio_frame1, text="(Recommended - Refreshes all data)",
                bg=colors['card'], fg=colors['text_secondary'],
                font=('Segoe UI', 9, 'italic')).pack(side="left", padx=(8, 0))

        radio_frame2 = tk.Frame(options_card_inner, bg=colors['card'])
        radio_frame2.pack(fill="x", pady=3)

        tk.Radiobutton(radio_frame2, text="Update only empty months",
                      variable=self.update_empty_var, value=True,
                      bg=colors['card'], fg=colors['text'],
                      font=('Segoe UI', 10),
                      activebackground=colors['card'],
                      selectcolor=colors['card'],
                      command=lambda: self.update_all_var.set(False)).pack(side="left")

        tk.Label(radio_frame2, text="(Only fills in missing data)",
                bg=colors['card'], fg=colors['text_secondary'],
                font=('Segoe UI', 9, 'italic')).pack(side="left", padx=(8, 0))

        # Status indicator
        status_frame = tk.Frame(content, bg='#E8F5E9', relief='flat')
        status_frame.pack(fill="x", pady=(0, 20))

        self.status_label = tk.Label(status_frame, text="✓ Ready to update",
                                     bg='#E8F5E9', fg=colors['success'],
                                     font=('Segoe UI', 10), pady=12)
        self.status_label.pack()

        # Action buttons
        btn_frame = tk.Frame(content, bg=colors['bg'])
        btn_frame.pack(pady=5)

        # Update button
        update_btn = tk.Button(btn_frame, text="Update Report",
                              command=lambda: self.run_update(dialog),
                              bg=colors['primary'], fg="white",
                              font=('Segoe UI', 11, 'bold'),
                              relief='flat', bd=0,
                              padx=35, pady=12,
                              cursor="hand2",
                              activebackground=colors['primary_hover'])
        update_btn.pack(side="left", padx=5)

        # Cancel button
        cancel_btn = tk.Button(btn_frame, text="Cancel",
                              command=dialog.destroy,
                              bg='#E0E0E0', fg=colors['text'],
                              font=('Segoe UI', 10),
                              relief='flat', bd=0,
                              padx=30, pady=12,
                              cursor="hand2",
                              activebackground='#D0D0D0')
        cancel_btn.pack(side="left", padx=5)
    
    def select_report(self, dialog):
        """Select existing report file"""
        path = filedialog.askopenfilename(
            title="Select Savings Report to Update",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            parent=dialog
        )

        if path:
            self.report_path = path
            filename = os.path.basename(path)
            # Truncate long filenames
            display_name = filename if len(filename) <= 50 else filename[:47] + "..."
            self.file_label.config(text=display_name, fg='#161616')
            self.status_label.config(
                text=f"✓ File selected: {filename}",
                bg='#E8F5E9',
                fg='#24A148'
            )

    def run_update(self, dialog):
        """Execute the report update"""

        if not self.report_path:
            messagebox.showerror("No File Selected",
                               "Please select a report file to update.",
                               parent=dialog)
            return

        # Show processing status
        self.status_label.config(
            text="⏳ Processing update... Please wait",
            bg='#FFF4E5',
            fg='#E87500'
        )
        dialog.update()

        try:
            success = self.update_report(self.update_all_var.get())

            if success:
                messagebox.showinfo(
                    "✓ Update Complete",
                    "Your report has been successfully updated!\n\n"
                    "The updated file has been saved with an '_Updated' suffix "
                    "and timestamp in the same directory.",
                    parent=dialog
                )
                dialog.destroy()
            else:
                self.status_label.config(
                    text="✗ Update failed - check error message above",
                    bg='#FFE8E8',
                    fg='#DA1E28'
                )

        except Exception as e:
            error_msg = str(e)
            messagebox.showerror("Update Error",
                               f"Failed to update report:\n\n{error_msg}",
                               parent=dialog)
            self.status_label.config(
                text="✗ Update failed - see error details",
                bg='#FFE8E8',
                fg='#DA1E28'
            )
    
    def update_report(self, update_all=True):
        """
        Main update logic - refreshes quantities and savings with latest data
        
        Args:
            update_all: If True, update all months. If False, only update empty months.
        """
        
        print(f"\n=== Starting Report Update ===")
        print(f"Report: {self.report_path}")
        print(f"Mode: {'Update ALL months' if update_all else 'Update EMPTY months only'}")
        
        # Load the existing report
        wb = load_workbook(self.report_path)
        
        if "Savings Report" not in wb.sheetnames:
            messagebox.showerror("Invalid Report", "This doesn't appear to be a valid savings report.")
            return False
        
        ws = wb["Savings Report"]
        
        # Load historical data
        hist = pd.read_excel(self.hist_path, dtype={"Product Code": str, "Vendor": str})
        hist = hist.rename(columns={
            "Product Code": "ProductCode",
            "Base RCVD QTY": "BaseRcvdQty",
            "Base Unit Cost": "UnitCost",
            "Receipt Date": "ReceiptDate"
        })
        hist = hist.dropna(subset=["Vendor", "ProductCode", "ReceiptDate", "UnitCost"]).copy()
        hist["ReceiptDate"] = pd.to_datetime(hist["ReceiptDate"], errors="coerce")
        hist = hist.dropna(subset=["ReceiptDate"])
        
        # Extract product information from SUMMARY section
        products = []
        summary_start = None
        
        for row_idx in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=1).value
            
            # Find SUMMARY header
            if cell_value and "SUMMARY" in str(cell_value).upper():
                summary_start = row_idx + 1  # Data starts after header
                break
        
        if not summary_start:
            messagebox.showerror("Parse Error", "Could not find SUMMARY section in report.")
            return False
        
        # Read products from summary (skip column header row)
        for row_idx in range(summary_start + 1, ws.max_row + 1):
            project_val = ws.cell(row=row_idx, column=1).value
            
            if not project_val or "QUANTITIES" in str(project_val).upper():
                break  # End of summary section
            
            product_code = ws.cell(row=row_idx, column=2).value
            compare_code = ws.cell(row=row_idx, column=3).value
            old_price = ws.cell(row=row_idx, column=5).value
            new_price = ws.cell(row=row_idx, column=6).value
            
            if product_code:
                products.append({
                    'product_code': str(product_code).strip(),
                    'compare_code': str(compare_code or product_code).strip(),
                    'old_price': float(old_price) if old_price else 0,
                    'new_price': float(new_price) if new_price else 0
                })
        
        print(f"Found {len(products)} products in report")
        
        # Find month columns and sections
        month_cols = {}  # {column_index: month_name}
        qty_section_start = None
        save_section_start = None

        print("\nScanning for sections...")
        for row_idx in range(1, ws.max_row + 1):
            # Check first few columns for section headers (handles merged cells)
            cell_text = None
            for col_check in range(1, min(5, ws.max_column + 1)):
                cell_value = ws.cell(row=row_idx, column=col_check).value
                if cell_value and str(cell_value).strip():
                    cell_text = str(cell_value).strip().upper()
                    break

            if cell_text:
                # More flexible pattern matching
                if "QUANTITIES" in cell_text or "QUANTITY" in cell_text:
                    qty_section_start = row_idx + 1  # Header is next row
                    print(f"  Found QUANTITIES section at row {row_idx}")
                elif ("SAVINGS" in cell_text and "REALIZED" in cell_text) or "REALIZED SAVINGS" in cell_text:
                    save_section_start = row_idx + 1
                    print(f"  Found SAVINGS section at row {row_idx}")

        if not qty_section_start or not save_section_start:
            print(f"  ⚠️ Section detection failed: qty_section_start={qty_section_start}, save_section_start={save_section_start}")
            messagebox.showerror("Parse Error",
                "Could not find QUANTITIES and SAVINGS sections.\n\n"
                "Please ensure your report has:\n"
                "- A 'QUANTITIES RECEIVED' section\n"
                "- A 'REALIZED SAVINGS' section")
            return False

        # Get month columns from quantities header
        if qty_section_start:
            for col_idx in range(1, ws.max_column + 1):
                header = ws.cell(row=qty_section_start, column=col_idx).value
                if header and str(header).strip() not in ["Product", "Total Qty", "Total", ""]:
                    # This is a month column (e.g., "JAN 25", "FEB 25")
                    month_cols[col_idx] = str(header).strip()

        print(f"Found {len(month_cols)} month columns: {list(month_cols.values())}")

        # Collect all unique months from ALL products first
        print("\nCollecting all months from historical data...")
        all_required_months = set()
        for product_info in products:
            product = product_info['product_code']
            product_hist = hist[hist["ProductCode"] == product].copy()

            if not product_hist.empty:
                product_hist["YearMonth"] = product_hist["ReceiptDate"].dt.to_period("M")
                monthly_data = product_hist.groupby("YearMonth").agg(
                    Qty=("BaseRcvdQty", "sum")
                ).reset_index()

                for period in monthly_data["YearMonth"]:
                    month_abbr = calendar.month_abbr[period.month].upper()
                    year_short = str(period.year)[-2:]
                    month_label = f"{month_abbr} {year_short}"
                    all_required_months.add(month_label)

        print(f"Found {len(all_required_months)} unique months in historical data")

        # Ensure all required month columns exist BEFORE processing
        if all_required_months:
            print("Ensuring all month columns exist...")
            for month_label in sorted(all_required_months):
                self.ensure_month_column(ws, qty_section_start, month_label)
                self.ensure_month_column(ws, save_section_start, month_label)

            # Rebuild month_cols mapping after all insertions
            month_cols = {}
            for col_idx in range(1, ws.max_column + 1):
                header = ws.cell(row=qty_section_start, column=col_idx).value
                if header and str(header).strip() not in ["Product", "Total Qty", "Total", ""]:
                    month_cols[col_idx] = str(header).strip()

            print(f"Final month columns: {list(month_cols.values())}")

        # Update each product's data
        updates_made = 0

        for product_info in products:
            product = product_info['product_code']
            compare = product_info['compare_code']
            old_price = product_info['old_price']
            new_price = product_info['new_price']
            
            print(f"\nUpdating: {product}")
            
            # Get product data from historical file
            product_hist = hist[hist["ProductCode"] == product].copy()
            
            if product_hist.empty:
                print(f"  ⚠️ No data found in historical file")
                continue
            
            # Group by month
            product_hist["YearMonth"] = product_hist["ReceiptDate"].dt.to_period("M")
            monthly_data = product_hist.groupby("YearMonth").agg(
                Qty=("BaseRcvdQty", "sum")
            ).reset_index()
            
            # Calculate savings per unit
            savings_per_unit = old_price - new_price

            # Update each month column
            update_empty = not update_all
            for col_idx, month_name in month_cols.items():
                # Parse month name (e.g., "JAN 25" -> 2025-01)
                try:
                    month_abbr, year_short = month_name.split()
                    month_num = list(calendar.month_abbr).index(month_abbr.upper())
                    year = 2000 + int(year_short)
                    target_period = pd.Period(f"{year}-{month_num:02d}", freq="M")
                except:
                    print(f"  ⚠️ Could not parse month: {month_name}")
                    continue

                # Get data for this month
                month_rows = monthly_data[monthly_data["YearMonth"] == target_period]

                if not month_rows.empty:
                    qty = month_rows.iloc[0]["Qty"]
                    savings = qty * savings_per_unit

                    # Update quantities section
                    qty_row = self.find_product_row(ws, product, qty_section_start)
                    if qty_row:
                        qty_cell = ws.cell(row=qty_row, column=col_idx)
                        self.safe_write(qty_cell, round(qty, 2), update_empty)
                        updates_made += 1

                    # Update savings section
                    save_row = self.find_product_row(ws, product, save_section_start)
                    if save_row:
                        save_cell = ws.cell(row=save_row, column=col_idx)
                        self.safe_write(save_cell, round(savings, 2), update_empty)

                    print(f"  Updated {month_name}: Qty={qty:.2f}, Savings=${savings:.2f}")
                else:
                    # No data for this month - set to 0
                    qty_row = self.find_product_row(ws, product, qty_section_start)
                    if qty_row:
                        qty_cell = ws.cell(row=qty_row, column=col_idx)
                        self.safe_write(qty_cell, 0, update_empty)

                    save_row = self.find_product_row(ws, product, save_section_start)
                    if save_row:
                        save_cell = ws.cell(row=save_row, column=col_idx)
                        self.safe_write(save_cell, 0, update_empty)
        
        # Recalculate totals for all columns
        print("\nRecalculating totals...")
        self.recalculate_totals(ws, month_cols, qty_section_start, save_section_start)
        
        # Update Monthly Summary sheet if it exists
        if "Monthly Summary" in wb.sheetnames:
            self.update_monthly_summary(wb, ws, month_cols, qty_section_start, save_section_start)
        
        # Save updated report
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        base_name = os.path.splitext(self.report_path)[0]
        output_path = f"{base_name}_Updated_{timestamp}.xlsx"
        
        wb.save(output_path)
        wb.close()
        
        print(f"\n✅ Update complete: {updates_made} cells updated")
        print(f"Saved to: {output_path}")
        
        return True
    
    def safe_write(self, cell, value, update_empty):
        """
        Safely write to a cell, skipping formulas and respecting update_empty flag

        Args:
            cell: The openpyxl cell object
            value: The value to write
            update_empty: If True, only write if cell is empty
        """
        # Skip formulas
        if cell.data_type == "f":
            return

        # Current content
        current = cell.value

        # Apply update_empty rule
        if update_empty and current not in [None, "", " ", 0, "0"]:
            return

        cell.value = value

    def ensure_month_column(self, ws, header_row, month_label):
        """
        Ensure the month column exists in the header_row.
        Return its 1-based column index.
        If missing:
          - Insert at correct chronological index
          - Clone formatting from nearest month column

        Args:
            ws: Worksheet object
            header_row: Row number containing month headers
            month_label: Month label to ensure (e.g., "JAN 25")

        Returns:
            Column index (1-based) for the month
        """
        # Parse target month
        try:
            month_abbr, year_short = month_label.split()
            target_month_num = list(calendar.month_abbr).index(month_abbr.upper())
            target_year = 2000 + int(year_short)
        except:
            print(f"  ⚠️ Could not parse month label: {month_label}")
            return None

        # Scan existing month columns
        month_columns = {}  # {col_idx: (month_num, year)}
        existing_col = None

        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(row=header_row, column=col_idx).value
            if header:
                header_str = str(header).strip()

                # Check if this is our target month
                if header_str == month_label:
                    existing_col = col_idx
                    break

                # Try to parse as month column
                try:
                    parts = header_str.split()
                    if len(parts) == 2:
                        m_abbr, y_short = parts
                        m_num = list(calendar.month_abbr).index(m_abbr.upper())
                        y = 2000 + int(y_short)
                        month_columns[col_idx] = (m_num, y)
                except:
                    pass

        # If month exists, return its column
        if existing_col:
            return existing_col

        # Month doesn't exist - need to insert it
        # Find correct chronological position
        insert_at = None
        template_col = None

        for col_idx in sorted(month_columns.keys()):
            m_num, y = month_columns[col_idx]

            # If current column is later than target, insert before it
            if (y > target_year) or (y == target_year and m_num > target_month_num):
                insert_at = col_idx
                template_col = col_idx
                break
            else:
                # This column is earlier, use as template
                template_col = col_idx

        # If no later column found, insert after last month column
        if insert_at is None and month_columns:
            last_col = max(month_columns.keys())
            insert_at = last_col + 1
            template_col = last_col
        elif insert_at is None:
            # No month columns found at all - insert after "Product" column
            insert_at = 2
            template_col = None

        # Insert the new column
        ws.insert_cols(insert_at)

        # Set header
        ws.cell(row=header_row, column=insert_at, value=month_label)

        # Clone formatting from template column if available
        if template_col:
            for row_idx in range(1, ws.max_row + 1):
                source_cell = ws.cell(row=row_idx, column=template_col)
                target_cell = ws.cell(row=row_idx, column=insert_at)

                target_cell.font = source_cell.font.copy()
                target_cell.fill = source_cell.fill.copy()
                target_cell.border = source_cell.border.copy()
                target_cell.alignment = source_cell.alignment.copy()
                target_cell.number_format = source_cell.number_format

        print(f"  ✓ Inserted new month column: {month_label} at position {insert_at}")
        return insert_at

    def find_product_row(self, ws, product_code, section_start):
        """Find the row for a specific product in a section"""
        for row_idx in range(section_start + 1, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=1).value

            # Stop at section end
            if not cell_value or "TOTAL" in str(cell_value).upper():
                break

            if str(cell_value).strip() == str(product_code).strip():
                return row_idx

        return None
    
    def recalculate_totals(self, ws, month_cols, qty_start, save_start):
        """Recalculate TOTAL rows for all month columns"""
        
        # Find TOTAL rows
        qty_total_row = None
        save_total_row = None
        
        for row_idx in range(qty_start, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=1).value
            if cell_value and "TOTAL" in str(cell_value).upper():
                qty_total_row = row_idx
                break
        
        for row_idx in range(save_start, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=1).value
            if cell_value and "TOTAL" in str(cell_value).upper():
                save_total_row = row_idx
                break
        
        # Calculate totals for each month column
        for col_idx in month_cols.keys():
            # Quantity totals
            if qty_total_row:
                total = 0
                for row_idx in range(qty_start + 1, qty_total_row):
                    val = ws.cell(row=row_idx, column=col_idx).value
                    if isinstance(val, (int, float)):
                        total += val
                ws.cell(row=qty_total_row, column=col_idx, value=round(total, 2))
            
            # Savings totals
            if save_total_row:
                total = 0
                for row_idx in range(save_start + 1, save_total_row):
                    val = ws.cell(row=row_idx, column=col_idx).value
                    if isinstance(val, (int, float)):
                        total += val
                ws.cell(row=save_total_row, column=col_idx, value=round(total, 2))
        
        # Also update "Total Qty" and "Total Savings" columns
        total_qty_col = None
        total_save_col = None
        
        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(row=qty_start, column=col_idx).value
            if header and "Total" in str(header):
                total_qty_col = col_idx
                break
        
        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(row=save_start, column=col_idx).value
            if header and "Total" in str(header):
                total_save_col = col_idx
                break
        
        # Recalculate row totals
        if total_qty_col:
            for row_idx in range(qty_start + 1, qty_total_row if qty_total_row else ws.max_row):
                row_total = 0
                for col_idx in month_cols.keys():
                    val = ws.cell(row=row_idx, column=col_idx).value
                    if isinstance(val, (int, float)):
                        row_total += val
                ws.cell(row=row_idx, column=total_qty_col, value=round(row_total, 2))
        
        if total_save_col:
            for row_idx in range(save_start + 1, save_total_row if save_total_row else ws.max_row):
                row_total = 0
                for col_idx in month_cols.keys():
                    val = ws.cell(row=row_idx, column=col_idx).value
                    if isinstance(val, (int, float)):
                        row_total += val
                ws.cell(row=row_idx, column=total_save_col, value=round(row_total, 2))
    
    def update_monthly_summary(self, wb, ws, month_cols, qty_start, save_start):
        """Update the Monthly Summary sheet with refreshed totals"""
        
        ws_summary = wb["Monthly Summary"]
        
        # Find TOTAL rows in main sheet
        qty_total_row = None
        save_total_row = None
        
        for row_idx in range(qty_start, ws.max_row + 1):
            if "TOTAL" in str(ws.cell(row=row_idx, column=1).value or "").upper():
                qty_total_row = row_idx
                break
        
        for row_idx in range(save_start, ws.max_row + 1):
            if "TOTAL" in str(ws.cell(row=row_idx, column=1).value or "").upper():
                save_total_row = row_idx
                break
        
        # Update each month in summary
        for col_idx, month_name in month_cols.items():
            # Find or create row for this month
            month_row = None
            for row_idx in range(2, ws_summary.max_row + 1):
                if ws_summary.cell(row=row_idx, column=1).value == month_name:
                    month_row = row_idx
                    break
            
            if not month_row:
                month_row = ws_summary.max_row + 1
                ws_summary.cell(row=month_row, column=1, value=month_name)
            
            # Get totals from main sheet
            qty_total = ws.cell(row=qty_total_row, column=col_idx).value if qty_total_row else 0
            save_total = ws.cell(row=save_total_row, column=col_idx).value if save_total_row else 0
            
            ws_summary.cell(row=month_row, column=2, value=round(qty_total or 0, 2))
            ws_summary.cell(row=month_row, column=3, value=round(save_total or 0, 2))


# Add this to your SavingsApp class in ui_main.py:

def add_update_report_button(self):
    """
    Add this method to your SavingsApp class and call it in create_menu_bar()
    
    Example:
    In your File Menu, add:
    file_menu.add_command(label="Update Existing Report...", command=self.show_report_updater)
    """
    pass

def show_report_updater(self):
    """Launch the report updater dialog"""
    updater = ReportUpdater(self.root, self.config)
    updater.show_updater_dialog()