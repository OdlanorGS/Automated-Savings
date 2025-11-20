"""
Costa Farms Sourcing Analysis Engine - With Dual Vendor Filtering
Focus on core functionality with separate target and baseline vendor filtering

Note: Vendor lists are passed as Python lists. The UI stores them as pipe-separated 
strings (" | ") to avoid conflicts with vendor names containing commas.
"""

import pandas as pd
import numpy as np
import os
import calendar
from datetime import datetime
from typing import List, Dict
from tkinter import messagebox
from config_manager import load_config, validate_hist_file

def normalize_vendor_name(vendor):
    """Normalize vendor names for consistent matching"""
    if pd.isna(vendor):
        return ""
    return str(vendor).strip().upper()

def run_savings_analysis(entries: List[Dict], project_name: str, track_months: int) -> str:
    """
    Analysis with dual vendor filtering:
    - Baseline vendors: for calculating prior year average cost
    - Target vendors: for tracking received quantities
    
    Args:
        entries: List of product entries with price changes and vendor filters
        project_name: Name of the project
        track_months: Number of months to track savings
    """
    
    config = load_config()
    if not validate_hist_file(config):
        return None
    
    hist_path = config["hist_file"]
    output_folder = config["output_folder"]
    
    if not output_folder:
        messagebox.showwarning("Missing Folder", 
                              "Please set an output folder in Settings before running.")
        return None
    
    print(f"\n{'='*60}")
    print(f"=== Starting Savings Analysis: {project_name} ===")
    print(f"{'='*60}\n")
    
    # Load historical data - OPTIMIZED
    try:
        print("Loading historical data... (this may take a moment for large files)")
        
        # Only load columns we need
        usecols = [
            "Sage & SAMII[Product Code]",
            "Sage & SAMII[Vendor Cleaned]",
            "Sage & SAMII[Receipt Date]",
            "[SumBase_RCVD_QTY]",
            "[SumBase_Unit_Cost]"
        ]
        
        hist = pd.read_excel(
            hist_path,
            usecols=usecols,
            dtype={
                "Sage & SAMII[Product Code]": str,
                "Sage & SAMII[Vendor Cleaned]": str
            }
        )
        
        print(f"✔ Loaded {len(hist):,} records from historical data")
        
    except Exception as e:
        messagebox.showerror("Data Error", f"Could not load historical data:\n{str(e)}")
        return None
    
    # Standardize column names
    hist = hist.rename(columns={
        "Sage & SAMII[Product Code]": "ProductCode",
        "[SumBase_RCVD_QTY]": "BaseRcvdQty", 
        "[SumBase_Unit_Cost]": "UnitCost",
        "Sage & SAMII[Receipt Date]": "ReceiptDate",
        "Sage & SAMII[Vendor Cleaned]": "Vendor"
    })
    
    # Clean data
    hist = hist.dropna(subset=["Vendor", "ProductCode", "ReceiptDate", "UnitCost"]).copy()
    hist["ReceiptDate"] = pd.to_datetime(hist["ReceiptDate"], errors="coerce")
    hist = hist.loc[hist["ReceiptDate"].notna()].copy()
    
    # OPTIMIZED: Convert to categorical for better performance
    hist["ProductCode"] = hist["ProductCode"].astype('category')
    hist["Vendor"] = hist["Vendor"].astype('category')
    hist["YearMonth"] = hist["ReceiptDate"].dt.to_period("M")
    
    # Add normalized vendor column for matching
    hist["VendorNormalized"] = hist["Vendor"].apply(normalize_vendor_name)
    
    print(f"✔ After cleaning: {len(hist):,} valid records")
    print(f"✔ Date range: {hist['ReceiptDate'].min()} to {hist['ReceiptDate'].max()}")
    print(f"✔ Unique products: {hist['ProductCode'].nunique():,}")
    print(f"✔ Unique vendors: {hist['Vendor'].nunique():,}")
    
    # Initialize result containers
    summary_rows = []
    qty_section = None
    save_section = None
    all_months = []
    
    # Process each product entry
    for idx, entry in enumerate(entries, 1):
        product = entry["product_code"].strip()
        compare_code = entry.get("compare_code", "").strip() or product
        new_price = float(entry["new_price"])
        effective_date = pd.Timestamp(entry["effective_date"])
        
        # Get vendor filters
        baseline_vendors = entry.get("baseline_vendors", "All")
        target_vendors = entry.get("target_vendors", "All")
        
        print(f"\n{'─'*60}")
        print(f"[{idx}/{len(entries)}] Processing: {product}")
        print(f"{'─'*60}")
        print(f"  Compare Code: {compare_code}")
        print(f"  New Price: ${new_price:.4f}")
        print(f"  Effective Date: {effective_date.date()}")
        
        # Normalize vendor filters
        baseline_vendors_normalized = None
        if baseline_vendors != "All" and isinstance(baseline_vendors, list):
            baseline_vendors_normalized = [normalize_vendor_name(v) for v in baseline_vendors]
            print(f"  Baseline Vendors (for cost): {baseline_vendors}")
        else:
            print(f"  Baseline Vendors (for cost): ALL")
        
        target_vendors_normalized = None
        if target_vendors != "All" and isinstance(target_vendors, list):
            target_vendors_normalized = [normalize_vendor_name(v) for v in target_vendors]
            print(f"  Target Vendors (for qty): {target_vendors}")
        else:
            print(f"  Target Vendors (for qty): ALL")
        
        # ==================== BASELINE CALCULATION ====================
        print(f"\n  [BASELINE] Looking for {compare_code}...")
        baseline_start = effective_date - pd.DateOffset(months=12)
        print(f"  [BASELINE] Date range: {baseline_start.date()} to {effective_date.date()}")
        
        baseline_df = hist[
            (hist["ProductCode"] == compare_code) &
            (hist["ReceiptDate"] < effective_date) &
            (hist["ReceiptDate"] >= baseline_start)
        ].copy()
        
        print(f"  [BASELINE] Found {len(baseline_df)} records before vendor filter")
        
        # Filter by baseline vendors if specified
        if baseline_vendors_normalized is not None:
            baseline_df = baseline_df[baseline_df["VendorNormalized"].isin(baseline_vendors_normalized)]
            print(f"  [BASELINE] After vendor filter: {len(baseline_df)} records")
            if len(baseline_df) > 0:
                print(f"  [BASELINE] Vendors matched: {baseline_df['Vendor'].unique().tolist()}")
        
        if baseline_df.empty:
            print(f"  ⚠️ [BASELINE] WARNING: No baseline data found!")
            print(f"     Checked product code: {compare_code}")
            if baseline_vendors_normalized:
                print(f"     Checked vendors: {baseline_vendors}")
                # Show what vendors ARE available for this product
                available = hist[hist["ProductCode"] == compare_code]["Vendor"].unique()
                print(f"     Available vendors in data: {list(available)[:10]}")
            continue
        
        # Simple average baseline price (Unit Cost only)
        # baseline_price = baseline_df["UnitCost"].mean() - not using anymore
        # Weighted average baseline price (weighted by quantity received)
        total_cost = (baseline_df["UnitCost"] * baseline_df["BaseRcvdQty"]).sum()
        total_qty = baseline_df["BaseRcvdQty"].sum()
        if total_qty > 0:
            baseline_price = total_cost / total_qty
        else:
            # Fallback to simple average if no quantity data
            baseline_price = baseline_df["UnitCost"].mean()
            print(f"  [BASELINE] ⚠️ Using simple average (no quantity data)")
        
        vendors_list = sorted(baseline_df["Vendor"].unique())
        vendors = ", ".join(vendors_list[:3])  # Limit to first 3 vendors for display
        
        print(f"  [BASELINE] ✓ Calculated baseline price: ${baseline_price:.4f}")
        print(f"  [BASELINE] ✓ From {len(vendors_list)} vendor(s): {vendors}")
        
        # ==================== REALIZED DATA ====================
        print(f"\n  [TARGET] Looking for {product}...")
        start_month = effective_date.replace(day=1)
        end_date = start_month + pd.offsets.MonthBegin(track_months)
        print(f"  [TARGET] Date range: {start_month.date()} to {end_date.date()}")
        
        months_list = [start_month + pd.offsets.MonthBegin(i) for i in range(track_months)]
        all_months.extend(months_list)
        
        realized = hist[
            (hist["ProductCode"] == product) &
            (hist["ReceiptDate"] >= start_month) &
            (hist["ReceiptDate"] < end_date)
        ].copy()
        
        print(f"  [TARGET] Found {len(realized)} records before vendor filter")
        
        # Filter by target vendors if specified
        if target_vendors_normalized is not None:
            realized = realized[realized["VendorNormalized"].isin(target_vendors_normalized)]
            print(f"  [TARGET] After vendor filter: {len(realized)} records")
            if len(realized) > 0:
                print(f"  [TARGET] Vendors matched: {realized['Vendor'].unique().tolist()}")
        
        if realized.empty:
            print(f"  ⚠️ [TARGET] WARNING: No receipts found!")
            print(f"     Checked product code: {product}")
            if target_vendors_normalized:
                print(f"     Checked vendors: {target_vendors}")
                # Show what vendors ARE available for this product
                available = hist[hist["ProductCode"] == product]["Vendor"].unique()
                print(f"     Available vendors in data: {list(available)[:10]}")
            continue
        
        # Calculate savings
        realized["RealizedSavings"] = realized["BaseRcvdQty"] * (baseline_price - new_price)
        
        # Group by month - ensure proper aggregation
        monthly = realized.groupby(realized["ReceiptDate"].dt.to_period("M")).agg(
            Qty=("BaseRcvdQty", "sum"),
            Savings=("RealizedSavings", "sum")
        ).reset_index()
        
        print(f"  [TARGET] ✓ Data found for {len(monthly)} month(s)")
        
        # Create month columns with short format (JAN 25, FEB 25, etc.)
        month_cols = []
        for m in months_list:
            month_abbr = calendar.month_abbr[m.month].upper()
            year_short = str(m.year)[2:]
            month_cols.append(f"{month_abbr} {year_short}")
        
        # Build quantity row with proper values
        qty_row = {"Product": product}
        for _, r in monthly.iterrows():
            month_abbr = calendar.month_abbr[r['ReceiptDate'].month].upper()
            year_short = str(r['ReceiptDate'].year)[2:]
            col_name = f"{month_abbr} {year_short}"
            if col_name in month_cols:
                qty_row[col_name] = round(r["Qty"], 2)
        
        # Fill missing months with 0
        for col in month_cols:
            if col not in qty_row:
                qty_row[col] = 0
        
        # Calculate total from actual values
        actual_total_qty = monthly["Qty"].sum()
        qty_row["Total Qty"] = round(actual_total_qty, 2)
        
        # Build savings row with proper values
        save_row = {"Product": product}
        for _, r in monthly.iterrows():
            month_abbr = calendar.month_abbr[r['ReceiptDate'].month].upper()
            year_short = str(r['ReceiptDate'].year)[2:]
            col_name = f"{month_abbr} {year_short}"
            if col_name in month_cols:
                save_row[col_name] = round(r["Savings"], 2)
        
        # Fill missing months with 0
        for col in month_cols:
            if col not in save_row:
                save_row[col] = 0
        
        # Calculate total from actual values
        actual_total_savings = monthly["Savings"].sum()
        save_row["Total Savings"] = round(actual_total_savings, 2)
        
        print(f"\n  ✅ SUCCESS: Total Qty: {actual_total_qty:.2f}, Total Savings: ${actual_total_savings:.2f}")
        
        # Create DataFrames
        qty_df = pd.DataFrame([qty_row], columns=["Product"] + month_cols + ["Total Qty"])
        save_df = pd.DataFrame([save_row], columns=["Product"] + month_cols + ["Total Savings"])
        
        # Append to sections
        qty_section = pd.concat([qty_section, qty_df], ignore_index=True) if qty_section is not None else qty_df
        save_section = pd.concat([save_section, save_df], ignore_index=True) if save_section is not None else save_df
        
        # Create summary row
        summary_rows.append({
            "Project": project_name,
            "Product Code": product,
            "Compare Code": compare_code,
            "Baseline Vendor(s)": vendors,
            "Target Vendor(s)": ", ".join(target_vendors[:3]) if isinstance(target_vendors, list) else "All",
            "Old Price (Avg)": round(baseline_price, 4),
            "New Price": new_price,
            "Price Change": round(new_price - baseline_price, 4),
            "Change %": round(((new_price - baseline_price) / baseline_price) * 100, 2),
            "Total Savings": round(save_row["Total Savings"], 2)
        })
    
    print(f"\n{'='*60}")
    if not summary_rows:
        print("❌ ANALYSIS FAILED: No valid data found")
        print("   Check the warnings above for each product")
        print(f"{'='*60}\n")
        messagebox.showwarning("No Data", 
            "No valid data found for analysis.\n\n"
            "Common issues:\n"
            "• Product codes don't match historical data\n"
            "• Selected vendors don't have data in date range\n"
            "• Date ranges don't contain any receipts\n\n"
            "Check the console output for detailed information.")
        return None
    
    print(f"✅ Successfully processed {len(summary_rows)} product(s)")
    print(f"{'='*60}\n")
    
    # Get unique months across all products
    all_months = sorted(set(all_months))
    all_month_names = []
    for m in all_months:
        month_abbr = calendar.month_abbr[m.month].upper()
        year_short = str(m.year)[2:]
        all_month_names.append(f"{month_abbr} {year_short}")
    
    # Ensure all products have all months
    for df in [qty_section, save_section]:
        if df is not None:
            for col in all_month_names:
                if col not in df.columns:
                    df[col] = 0
            # Reorder columns
            df = df[["Product"] + all_month_names + [df.columns[-1]]]
    
    # Add totals row with correct summation
    if qty_section is not None:
        totals_row = {"Product": "TOTAL"}
        for col in all_month_names:
            if col in qty_section.columns:
                col_sum = qty_section[col].apply(lambda x: x if isinstance(x, (int, float)) else 0).sum()
                totals_row[col] = round(col_sum, 2)
            else:
                totals_row[col] = 0
        
        totals_row["Total Qty"] = round(
            qty_section["Total Qty"].apply(lambda x: x if isinstance(x, (int, float)) else 0).sum(), 2
        )
        
        qty_section = pd.concat([qty_section, pd.DataFrame([totals_row])], ignore_index=True)
    
    if save_section is not None:
        totals_row = {"Product": "TOTAL"}
        for col in all_month_names:
            if col in save_section.columns:
                col_sum = save_section[col].apply(lambda x: x if isinstance(x, (int, float)) else 0).sum()
                totals_row[col] = round(col_sum, 2)
            else:
                totals_row[col] = 0
        
        totals_row["Total Savings"] = round(
            save_section["Total Savings"].apply(lambda x: x if isinstance(x, (int, float)) else 0).sum(), 2
        )
        
        save_section = pd.concat([save_section, pd.DataFrame([totals_row])], ignore_index=True)
    
    # Create summary DataFrame
    summary_df = pd.DataFrame(summary_rows)
    
    # Generate output file with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    output_file = f"Savings_Analysis_{project_name}_{timestamp}.xlsx"
    output_path = os.path.join(output_folder, output_file)
    
    # Write to Excel with clean structure
    try:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            sheet_name = "Savings Report"
            current_row = 1
            
            # Track section positions
            section_positions = {}
            
            # Write SUMMARY section
            section_positions['summary'] = current_row
            summary_df.to_excel(writer, sheet_name=sheet_name, 
                               startrow=current_row, index=False)
            current_row += len(summary_df) + 1 + 3
            
            # Write QUANTITIES section
            if qty_section is not None and not qty_section.empty:
                section_positions['quantities'] = current_row
                qty_section.to_excel(writer, sheet_name=sheet_name, 
                                    startrow=current_row, index=False)
                current_row += len(qty_section) + 1 + 3
            
            # Write SAVINGS section
            if save_section is not None and not save_section.empty:
                section_positions['savings'] = current_row
                save_section.to_excel(writer, sheet_name=sheet_name,
                                     startrow=current_row, index=False)
            
            # Add styled section headers
            ws = writer.sheets[sheet_name]
            
            def add_section_header(title: str, row: int):
                """Insert and style section header above data table"""
                ws.insert_rows(row)
                ws.merge_cells(start_row=row, start_column=1, 
                              end_row=row, end_column=ws.max_column or 10)
                
                header_cell = ws.cell(row=row, column=1)
                header_cell.value = title
                header_cell.font = Font(bold=True, size=13, color="2D5016")
                header_cell.alignment = Alignment(horizontal="center", vertical="center")
                header_cell.fill = PatternFill(fill_type="solid", 
                                              start_color="E8F5E8",
                                              end_color="E8F5E8")
            
            # Insert headers from bottom to top
            if 'savings' in section_positions:
                add_section_header("REALIZED SAVINGS", section_positions['savings'])
            
            if 'quantities' in section_positions:
                add_section_header("QUANTITIES RECEIVED", section_positions['quantities'])
            
            add_section_header("SUMMARY", section_positions['summary'])
            
            # Create Monthly Summary sheet
            monthly_summary = []
            for col in all_month_names:
                month_qty = 0
                month_savings = 0
                
                if qty_section is not None and col in qty_section.columns:
                    for idx in range(len(qty_section) - 1):
                        val = qty_section.iloc[idx][col]
                        if isinstance(val, (int, float)):
                            month_qty += val
                
                if save_section is not None and col in save_section.columns:
                    for idx in range(len(save_section) - 1):
                        val = save_section.iloc[idx][col]
                        if isinstance(val, (int, float)):
                            month_savings += val
                
                monthly_summary.append({
                    "Month": col,
                    "Total Qty": round(month_qty, 2),
                    "Total Savings": round(month_savings, 2)
                })
            
            # Add grand totals row
            total_qty_all = sum(row["Total Qty"] for row in monthly_summary)
            total_savings_all = sum(row["Total Savings"] for row in monthly_summary)
            monthly_summary.append({
                "Month": "GRAND TOTAL",
                "Total Qty": round(total_qty_all, 2),
                "Total Savings": round(total_savings_all, 2)
            })
            
            # Write Monthly Summary sheet
            monthly_df = pd.DataFrame(monthly_summary)
            monthly_df.to_excel(writer, sheet_name="Monthly Summary", index=False)
        
        print(f"✅ Report saved: {output_path}\n")
        
        total_savings = sum(r["Total Savings"] for r in summary_rows)
        messagebox.showinfo("Analysis Complete", 
                           f"Report generated successfully!\n\n"
                           f"File: {output_file}\n"
                           f"Total Savings: ${total_savings:,.2f}")
        
        return output_path
        
    except Exception as e:
        messagebox.showerror("Export Error", f"Failed to create Excel file:\n{str(e)}")
        return None